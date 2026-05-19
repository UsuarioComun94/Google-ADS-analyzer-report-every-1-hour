from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

MASTER_PATH = Path(r"D:\Proyectos\hma-system\historico\HMA_Master.xlsx")
SHEET_NAME = "metric_crosses"

HEADER_MAP = {
    "timestamp": "fecha_hora",
    "client_number": "cliente",
    "cross_id": "id_cruce",
    "pattern": "patron",
    "severity": "severidad",
    "confidence": "confianza",
    "sample_status": "estado_muestra",
    "comparison_gap_hours": "brecha_comparacion_horas",
    "comparison_quality": "calidad_comparacion",
    "likely_bottleneck": "diagnostico_probable",
    "evidence": "evidencia",
    "recommended_action": "accion_recomendada",
    "blocked_action": "accion_bloqueada",
    "required_checks": "controles_requeridos",
    "source_metrics": "metricas_fuente",
}

VALUE_MAP = {
    "critical": "crítico",
    "high": "alto",
    "medium/high": "media/alta",
    "medium": "media",
    "low": "baja",
    "warning": "advertencia",
    "opportunity": "oportunidad",
    "info": "informativo",
    "valid": "válido",
    "unknown": "desconocido",
    "low_clicks": "muestra baja de clicks",
    "low_conversions": "muestra baja de conversiones",
    "clean_hourly_comparison": "comparación horaria limpia",
    "degraded_missing_hours": "comparación degradada por horas faltantes",
}

PATTERN_MAP = {
    "CPA up + ROAS down": "CPA sube + ROAS baja",
    "ROAS down + revenue down": "ROAS baja + ingresos bajan",
    "CTR up + CVR down": "CTR sube + CVR baja",
    "CVR down + CPA up": "CVR baja + CPA sube",
    "CTR down + CPC up": "CTR baja + CPC sube",
    "Spend up + conversions not following": "Inversión sube + conversiones no acompañan",
    "Clicks up + conversions flat/down": "Clicks suben + conversiones planas/bajan",
    "CPC down + CVR stable/up": "CPC baja + CVR estable/sube",
    "ROAS up + CPA down + conversions up": "ROAS sube + CPA baja + conversiones suben",
    "High clicks + zero conversions": "Clicks altos + conversiones en cero",
    "Comparison gap > 1 hour": "Brecha de comparación mayor a 1 hora",
    "Low sample / insufficient statistical base": "Muestra baja / base estadística insuficiente",
}

TOKEN_MAP = {
    "profitability": "rentabilidad",
    "conversion_quality": "calidad de conversión",
    "budget_efficiency": "eficiencia de presupuesto",
    "revenue_quality": "calidad de ingresos",
    "conversion_value": "valor de conversión",
    "commercial_value": "valor comercial",
    "landing": "landing",
    "offer": "oferta",
    "form": "formulario",
    "message_match": "congruencia anuncio-página",
    "tracking": "tracking",
    "traffic_quality": "calidad de tráfico",
    "creative": "creatividad",
    "relevance": "relevancia",
    "auction_pressure": "presión de subasta",
    "keyword_match": "match de keywords",
    "audience": "audiencia",
    "budget_pacing": "ritmo de presupuesto",
    "conversion_efficiency": "eficiencia de conversión",
    "efficient_traffic": "tráfico eficiente",
    "auction_improvement": "mejora de subasta",
    "scale_candidate": "candidato a escala",
    "efficient_growth": "crecimiento eficiente",
    "data_continuity": "continuidad de datos",
    "missing_hours": "horas faltantes",
    "sample_size": "tamaño de muestra",
    "decision_quality": "calidad de decisión",
    "tracking_health": "salud del tracking",
    "sample_status": "estado de muestra",
    "comparison_gap": "brecha de comparación",
    "rolling_average": "promedio móvil",
    "auction_context": "contexto de subasta",
    "revenue_validation": "validación de ingresos",
    "required_manual_validation": "validación manual requerida",
    "timestamp": "fecha/hora",
    "compared_against_timestamp": "fecha/hora comparada",
    "spend": "inversión",
    "clicks": "clicks",
    "conversions": "conversiones",
    "revenue": "ingresos",
    "ctr": "CTR",
    "cpc": "CPC",
    "cvr": "CVR",
    "cpa": "CPA",
    "roas": "ROAS",
}

def clean(v):
    return "" if v is None else str(v).strip()

def lower(v):
    return clean(v).lower()

def translate_pipe_tokens(value):
    text = clean(value)
    if not text:
        return value

    parts = [p.strip() for p in text.split("|")]
    translated = [TOKEN_MAP.get(p, p) for p in parts]
    return " | ".join(translated)

wb = load_workbook(MASTER_PATH)

if SHEET_NAME not in wb.sheetnames:
    raise SystemExit("No existe metric_crosses.")

ws = wb[SHEET_NAME]

# Traducir encabezados sin cambiar orden ni borrar columnas.
for cell in ws[1]:
    key = lower(cell.value)
    if key in HEADER_MAP:
        cell.value = HEADER_MAP[key]

# Detectar columnas por encabezado ya traducido.
headers = {}
for cell in ws[1]:
    headers[lower(cell.value)] = cell.column

pattern_col = headers.get("patron")
severity_col = headers.get("severidad")
confidence_col = headers.get("confianza")
sample_col = headers.get("estado_muestra")
quality_col = headers.get("calidad_comparacion")
bottleneck_col = headers.get("diagnostico_probable")
checks_col = headers.get("controles_requeridos")
metrics_col = headers.get("metricas_fuente")
id_col = headers.get("id_cruce")

for row in range(2, ws.max_row + 1):
    if pattern_col:
        cell = ws.cell(row=row, column=pattern_col)
        cell.value = PATTERN_MAP.get(clean(cell.value), cell.value)

    for col in [severity_col, confidence_col, sample_col, quality_col]:
        if col:
            cell = ws.cell(row=row, column=col)
            cell.value = VALUE_MAP.get(lower(cell.value), cell.value)

    for col in [bottleneck_col, checks_col, metrics_col]:
        if col:
            cell = ws.cell(row=row, column=col)
            cell.value = translate_pipe_tokens(cell.value)

# Mantener id_cruce/cross_id oculto: es t?cnico, no operativo.
if id_col:
    letter = ws.cell(row=1, column=id_col).column_letter
    ws.column_dimensions[letter].hidden = True
    ws.column_dimensions[letter].width = 34

# Estilo sobrio solo para metric_crosses.
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
thin = Border(bottom=Side(style="thin", color="E8EEF7"), right=Side(style="thin", color="E8EEF7"))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

severity_colors = {
    "crítico": ("FEE2E2", "991B1B"),
    "alto": ("FEF3C7", "92400E"),
    "advertencia": ("E0F2FE", "075985"),
    "oportunidad": ("DCFCE7", "166534"),
    "informativo": ("F1F5F9", "334155"),
}

if severity_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=severity_col)
        sev = lower(cell.value)
        fill, font = severity_colors.get(sev, ("FFFFFF", "111827"))
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=10, bold=True, color=font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 58
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin

ws.sheet_properties.tabColor = "B71C1C"

wb.save(MASTER_PATH)

print("metric_crosses traducida a español sin tocar otras hojas.")
