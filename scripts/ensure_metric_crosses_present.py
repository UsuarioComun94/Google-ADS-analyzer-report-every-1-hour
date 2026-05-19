from __future__ import annotations

import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"D:\Proyectos\hma-system")
MASTER = BASE / "historico" / "HMA_Master.xlsx"
BACKUP_DIR = BASE / "historico" / "_metric_crosses_guard_backups"
BUILD_SCRIPT = BASE / "scripts" / "build_metric_crosses.py"
TRANSLATE_SCRIPT = BASE / "scripts" / "translate_metric_crosses_only.py"

BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def main() -> int:
    if not MASTER.exists():
        print(f"No existe HMA_Master.xlsx: {MASTER}")
        return 1

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup = BACKUP_DIR / f"HMA_Master_before_metric_crosses_guard_{stamp}.xlsx"
    shutil.copy2(MASTER, backup)
    print(f"Backup guard creado: {backup}")

    wb = load_workbook(MASTER)

    if "metric_crosses" not in wb.sheetnames:
        wb.close()
        print("metric_crosses no existe. Reconstruyendo desde build_metric_crosses.py...")

        if not BUILD_SCRIPT.exists():
            print(f"No existe {BUILD_SCRIPT}")
            return 2

        result = subprocess.run([sys.executable, str(BUILD_SCRIPT)], cwd=str(BASE), text=True)

        if result.returncode != 0:
            print(f"build_metric_crosses.py fall? con c?digo {result.returncode}")
            return result.returncode

    else:
        wb.close()
        print("metric_crosses existe. Se verifica visibilidad.")

    if TRANSLATE_SCRIPT.exists():
        result = subprocess.run([sys.executable, str(TRANSLATE_SCRIPT)], cwd=str(BASE), text=True)
        if result.returncode != 0:
            print(f"Advertencia: translate_metric_crosses_only.py fall? con c?digo {result.returncode}. Se contin?a.")

    wb = load_workbook(MASTER)

    if "metric_crosses" not in wb.sheetnames:
        print("ERROR: metric_crosses sigue sin existir despu?s del guard.")
        return 3

    ws = wb["metric_crosses"]
    ws.sheet_state = "visible"
    ws.sheet_properties.tabColor = "B71C1C"

    if wb.views:
        wb.views[0].showSheetTabs = True

    # Mostrar todas las columnas primero.
    for col_letter in list(ws.column_dimensions.keys()):
        ws.column_dimensions[col_letter].hidden = False

    hidden = []
    pattern_visible = []

    for cell in ws[1]:
        header = "" if cell.value is None else str(cell.value).strip().lower()

        if header in {"id_cruce", "cross_id"}:
            ws.column_dimensions[cell.column_letter].hidden = True
            ws.column_dimensions[cell.column_letter].width = 34
            hidden.append(f"{cell.column_letter} ({header})")

        if header in {"patron", "patr?n", "pattern"}:
            ws.column_dimensions[cell.column_letter].hidden = False
            ws.column_dimensions[cell.column_letter].width = 36
            pattern_visible.append(f"{cell.column_letter} ({header})")

    wb.save(MASTER)

    print("metric_crosses asegurada.")
    print("Columnas t?cnicas ocultas:", hidden)
    print("Patr?n visible:", pattern_visible)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
