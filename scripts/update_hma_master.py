from pathlib import Path
import json
import math
import re
import hashlib
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule


BASE_DIR = Path(__file__).resolve().parent.parent
DOWNLOADS_DIR = BASE_DIR / "downloads"
HISTORY_DIR = BASE_DIR / "historico"
MASTER_FILE = HISTORY_DIR / "HMA_Master.xlsx"
FINGERPRINT_FILE = HISTORY_DIR / "HMA_Master.fingerprint.json"

HISTORY_DIR.mkdir(exist_ok=True)

# Criterio comercial base para presupuestos diarios altos.
CHANGE_NOISE_PCT = 2.0
CHANGE_MODERATE_PCT = 5.0
CHANGE_CRITICAL_PCT = 10.0

DEFAULT_DAILY_BUDGET_USD = 30000.0
MAX_SCALE_UP_PCT = 10.0
MAX_SCALE_DOWN_PCT = 10.0


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, float) and math.isnan(value):
            return default
        return float(value)
    except Exception:
        return default


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        if isinstance(value, float) and math.isnan(value):
            return default
        return int(float(value))
    except Exception:
        return default


def pct_change(current: float, previous: float) -> float:
    if previous == 0:
        return 0.0
    return ((current - previous) / previous) * 100


def direction(change_pct: float) -> str:
    if abs(change_pct) < CHANGE_NOISE_PCT:
        return "estable"
    if change_pct > 0:
        return "subió"
    return "bajó"


def relevance(change_pct: float) -> str:
    abs_change = abs(change_pct)

    if abs_change < CHANGE_NOISE_PCT:
        return "ruido normal"
    if abs_change < CHANGE_MODERATE_PCT:
        return "leve"
    if abs_change < CHANGE_CRITICAL_PCT:
        return "moderada"
    return "crítica"


def metric_business_interpretation(metric: str, change_pct: float) -> str:
    metric = metric.lower()

    if abs(change_pct) < CHANGE_NOISE_PCT:
        return "Variación menor al umbral. No tomar decisión por esta señal aislada."

    if metric == "cpa":
        return (
            "CPA subió: eficiencia deteriorada."
            if change_pct > 0
            else "CPA bajó: eficiencia mejorada."
        )

    if metric == "roas":
        return (
            "ROAS subió: rentabilidad mejorada."
            if change_pct > 0
            else "ROAS bajó: rentabilidad deteriorada."
        )

    if metric == "ctr":
        return (
            "CTR subió: mejor respuesta creativa o mayor relevancia."
            if change_pct > 0
            else "CTR bajó: posible problema de títulos, descripciones, imágenes o relevancia."
        )

    if metric == "cvr":
        return (
            "CVR subió: mejor calidad de tráfico o conversión."
            if change_pct > 0
            else "CVR bajó: posible problema de intención, landing, oferta o tracking."
        )

    if metric == "spend":
        return (
            "Gasto subió: revisar si el incremento está acompañado por conversiones."
            if change_pct > 0
            else "Gasto bajó: revisar pacing y entrega."
        )

    if metric == "clicks":
        return (
            "Clics subieron: revisar si las conversiones acompañan."
            if change_pct > 0
            else "Clics bajaron: revisar tráfico, anuncios o pérdida de volumen."
        )

    if metric == "conversions":
        return (
            "Conversiones subieron: señal positiva si CPA/ROAS acompañan."
            if change_pct > 0
            else "Conversiones bajaron: revisar tráfico, funnel o tracking."
        )

    return "Cambio registrado."


def load_json_safely(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def extract_run_id_from_path(path: Path) -> str:
    """
    Extrae el GitHub Run ID desde carpetas tipo:
    21-30_run-25705041490
    Si no existe, devuelve vacío.
    """
    match = re.search(r"run-(\d+)", str(path))
    return match.group(1) if match else ""


def find_artifact_runs() -> list[dict]:
    """
    Busca artifacts descargados en downloads/ y deduplica para que:
    1 GitHub Run ID = 1 registro horario válido.

    Si no existe Run ID en carpetas viejas, deduplica por:
    client_number + latest_timestamp.
    """
    raw_runs = []

    if not DOWNLOADS_DIR.exists():
        return raw_runs

    summary_files = list(DOWNLOADS_DIR.rglob("latest_summary.json"))

    for summary_path in summary_files:
        exports_dir = summary_path.parent
        metrics_path = exports_dir / "latest_metrics.csv"
        connection_path = exports_dir / "connection_status.json"

        if not metrics_path.exists():
            continue

        artifact_root = exports_dir.parent
        reports_dir = artifact_root / "reports"

        named_report = ""
        latest_report = ""

        if reports_dir.exists():
            latest_candidate = reports_dir / "latest_hourly_report.md"
            if latest_candidate.exists():
                latest_report = str(latest_candidate)

            named_reports = [
                p for p in reports_dir.glob("*.md")
                if p.name != "latest_hourly_report.md"
            ]
            if named_reports:
                named_report = str(named_reports[0])

        run_key = str(artifact_root.relative_to(DOWNLOADS_DIR))
        run_id = extract_run_id_from_path(artifact_root)
        summary = load_json_safely(summary_path)

        client_number = summary.get("client_number", "")
        latest_timestamp = summary.get("latest_timestamp", "")

        if run_id:
            dedupe_key = f"run_id::{run_id}"
        else:
            dedupe_key = f"timestamp::{client_number}::{latest_timestamp}"

        raw_runs.append(
            {
                "dedupe_key": dedupe_key,
                "run_id": run_id,
                "run_key": run_key,
                "artifact_root": artifact_root,
                "summary_path": summary_path,
                "metrics_path": metrics_path,
                "connection_path": connection_path if connection_path.exists() else None,
                "latest_report": latest_report,
                "named_report": named_report,
                "client_number": client_number,
                "latest_timestamp": latest_timestamp,
            }
        )

    unique = {}
    for run in sorted(raw_runs, key=lambda item: str(item["artifact_root"])):
        if run["dedupe_key"] not in unique:
            unique[run["dedupe_key"]] = run

    return list(unique.values())


def load_summary(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def build_hourly_summary(runs: list[dict]) -> pd.DataFrame:
    rows = []

    for run in runs:
        summary = load_summary(run["summary_path"])
        latest = summary.get("latest", {})
        thresholds = summary.get("thresholds", {})
        changes = summary.get("changes", {})
        alerts = summary.get("alerts", [])
        connection_status = summary.get("connection_status", {})
        business_context = summary.get("business_context", {})

        daily_budget_usd = safe_float(
            business_context.get("daily_budget_usd"),
            DEFAULT_DAILY_BUDGET_USD,
        )
        hourly_budget_usd = safe_float(
            business_context.get("hourly_budget_usd"),
            daily_budget_usd / 24,
        )

        rows.append(
            {
                "dedupe_key": run["dedupe_key"],
                "run_id": run["run_id"],
                "run_key": run["run_key"],
                "client_number": summary.get("client_number", ""),
                "timestamp": summary.get("latest_timestamp", ""),
                "timezone": summary.get("report_timezone", ""),
                "data_source": summary.get("source", ""),
                "requested_data_source": summary.get("requested_data_source", ""),
                "health_status": summary.get("health_status", ""),
                "daily_budget_usd": daily_budget_usd,
                "hourly_budget_usd": hourly_budget_usd,
                "spend": safe_float(latest.get("spend")),
                "impressions": safe_int(latest.get("impressions")),
                "clicks": safe_int(latest.get("clicks")),
                "ctr": safe_float(latest.get("ctr")),
                "cpc": safe_float(latest.get("cpc")),
                "conversions": safe_int(latest.get("conversions")),
                "cvr": safe_float(latest.get("cvr")),
                "cpa": safe_float(latest.get("cpa")),
                "revenue": safe_float(latest.get("revenue")),
                "roas": safe_float(latest.get("roas")),
                "alerts_count": len(alerts),
                "cpa_threshold": safe_float(thresholds.get("cpa_threshold")),
                "roas_threshold": safe_float(thresholds.get("roas_threshold")),
                "ctr_threshold": safe_float(thresholds.get("ctr_threshold")),
                "spend_change_pct_report": safe_float(changes.get("spend_change_pct")),
                "cpa_change_pct_report": safe_float(changes.get("cpa_change_pct")),
                "roas_change_pct_report": safe_float(changes.get("roas_change_pct")),
                "google_ads_ready": connection_status.get("google_ads_ready", False),
                "meta_ads_ready": connection_status.get("meta_ads_ready", False),
                "named_report_path": run["named_report"],
                "latest_report_path": run["latest_report"],
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df["timestamp_dt"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = (
        df.sort_values(["timestamp_dt", "run_id", "run_key"])
        .drop_duplicates(subset=["dedupe_key"], keep="last")
        .drop_duplicates(subset=["client_number", "timestamp"], keep="last")
    )

    return df.sort_values("timestamp_dt")


def build_campaign_metrics(runs: list[dict]) -> pd.DataFrame:
    frames = []

    for run in runs:
        try:
            summary = load_summary(run["summary_path"])
            latest_timestamp = summary.get("latest_timestamp", "")
            client_number = summary.get("client_number", "")

            df = pd.read_csv(run["metrics_path"])
            df["dedupe_key"] = run["dedupe_key"]
            df["run_id"] = run["run_id"]
            df["run_key"] = run["run_key"]
            df["client_number"] = client_number
            df["summary_timestamp"] = latest_timestamp

            frames.append(df)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()

    df_all = pd.concat(frames, ignore_index=True)
    df_all["summary_timestamp_dt"] = pd.to_datetime(
        df_all["summary_timestamp"],
        errors="coerce",
    )

    dedupe_cols = [
        "dedupe_key",
        "platform",
        "account_name",
        "campaign_name",
        "creative_group",
    ]
    existing_cols = [col for col in dedupe_cols if col in df_all.columns]

    if existing_cols:
        df_all = df_all.drop_duplicates(subset=existing_cols, keep="last")

    return df_all.sort_values(["summary_timestamp_dt", "platform", "campaign_name"])


def build_metric_comparison(hourly_df: pd.DataFrame) -> pd.DataFrame:
    if hourly_df.empty or len(hourly_df) < 2:
        return pd.DataFrame()

    hourly_df = hourly_df.sort_values("timestamp_dt").reset_index(drop=True)

    metrics = [
        "spend",
        "impressions",
        "clicks",
        "ctr",
        "cpc",
        "conversions",
        "cvr",
        "cpa",
        "revenue",
        "roas",
    ]

    rows = []

    for idx in range(1, len(hourly_df)):
        current = hourly_df.iloc[idx]
        previous = hourly_df.iloc[idx - 1]

        for metric in metrics:
            current_value = safe_float(current.get(metric))
            previous_value = safe_float(previous.get(metric))
            change = pct_change(current_value, previous_value)

            rows.append(
                {
                    "timestamp": current["timestamp"],
                    "client_number": current["client_number"],
                    "run_id": current.get("run_id", ""),
                    "compared_against_timestamp": previous["timestamp"],
                    "compared_against_run_id": previous.get("run_id", ""),
                    "metric": metric,
                    "current_value": current_value,
                    "previous_value": previous_value,
                    "change_pct": change,
                    "direction": direction(change),
                    "relevance": relevance(change),
                    "commercial_interpretation": metric_business_interpretation(
                        metric,
                        change,
                    ),
                }
            )

    return pd.DataFrame(rows)


def get_last_change(comparison_df: pd.DataFrame, metric: str) -> float:
    if comparison_df.empty:
        return 0.0

    metric_df = comparison_df[comparison_df["metric"] == metric]
    if metric_df.empty:
        return 0.0

    return safe_float(metric_df.iloc[-1]["change_pct"])


def confidence_from_signals(
    negative_signals: int,
    positive_signals: int,
    history_len: int,
) -> str:
    if history_len < 2:
        return "BAJA"

    if negative_signals >= 3 or positive_signals >= 3:
        return "ALTA" if history_len >= 3 else "MEDIA"

    if negative_signals >= 2 or positive_signals >= 2:
        return "MEDIA"

    return "BAJA"


def build_recommendations(hourly_df: pd.DataFrame, comparison_df: pd.DataFrame) -> pd.DataFrame:
    if hourly_df.empty:
        return pd.DataFrame()

    latest = hourly_df.sort_values("timestamp_dt").iloc[-1]
    history_len = len(hourly_df)

    cpa_change = get_last_change(comparison_df, "cpa")
    roas_change = get_last_change(comparison_df, "roas")
    ctr_change = get_last_change(comparison_df, "ctr")
    cvr_change = get_last_change(comparison_df, "cvr")
    spend_change = get_last_change(comparison_df, "spend")
    clicks_change = get_last_change(comparison_df, "clicks")
    conversions_change = get_last_change(comparison_df, "conversions")

    cpa = safe_float(latest.get("cpa"))
    roas = safe_float(latest.get("roas"))
    ctr = safe_float(latest.get("ctr"))
    conversions = safe_int(latest.get("conversions"))
    clicks = safe_int(latest.get("clicks"))
    spend = safe_float(latest.get("spend"))

    cpa_threshold = safe_float(latest.get("cpa_threshold"), 25)
    roas_threshold = safe_float(latest.get("roas_threshold"), 1.5)
    ctr_threshold = safe_float(latest.get("ctr_threshold"), 1.0)

    daily_budget_usd = safe_float(
        latest.get("daily_budget_usd"),
        DEFAULT_DAILY_BUDGET_USD,
    )
    hourly_budget_usd = safe_float(
        latest.get("hourly_budget_usd"),
        daily_budget_usd / 24,
    )

    rows = []

    def add(
        level: str,
        action: str,
        motive: str,
        area: str,
        confidence: str,
        human_review: bool = True,
    ):
        rows.append(
            {
                "timestamp": latest.get("timestamp", ""),
                "client_number": latest.get("client_number", ""),
                "run_id": latest.get("run_id", ""),
                "level": level,
                "area": area,
                "recommended_action": action,
                "motive": motive,
                "confidence": confidence,
                "requires_human_review": human_review,
                "daily_budget_usd_reference": daily_budget_usd,
                "hourly_budget_usd_reference": hourly_budget_usd,
            }
        )

    negative_signals = 0
    positive_signals = 0

    if cpa_change > CHANGE_MODERATE_PCT:
        negative_signals += 1
    if roas_change < -CHANGE_MODERATE_PCT:
        negative_signals += 1
    if ctr_change < -CHANGE_MODERATE_PCT:
        negative_signals += 1
    if cvr_change < -CHANGE_MODERATE_PCT:
        negative_signals += 1
    if conversions_change < -CHANGE_MODERATE_PCT:
        negative_signals += 1

    if cpa_change < -CHANGE_MODERATE_PCT:
        positive_signals += 1
    if roas_change > CHANGE_MODERATE_PCT:
        positive_signals += 1
    if conversions_change > CHANGE_MODERATE_PCT:
        positive_signals += 1
    if cvr_change > CHANGE_MODERATE_PCT:
        positive_signals += 1

    confidence = confidence_from_signals(
        negative_signals=negative_signals,
        positive_signals=positive_signals,
        history_len=history_len,
    )

    if spend > 0 and clicks > 100 and conversions == 0:
        add(
            "CRÍTICO",
            "No cambiar presupuesto hasta validar tracking. Revisar pixel, conversion API, tags, evento de conversión, landing y caída del sitio.",
            "Hay gasto y volumen de clics, pero conversiones en cero.",
            "tracking/funnel",
            "MEDIA" if history_len >= 2 else "BAJA",
        )

    if cpa > cpa_threshold and cpa_change > CHANGE_CRITICAL_PCT and roas_change < -CHANGE_MODERATE_PCT:
        add(
            "ALTO",
            f"Frenar escalado o reducir presupuesto hasta {MAX_SCALE_DOWN_PCT:.0f}% con revisión humana.",
            f"CPA por encima del umbral y subiendo {cpa_change:.1f}%, con ROAS cayendo {roas_change:.1f}%.",
            "presupuesto/eficiencia",
            confidence,
        )

    elif cpa > cpa_threshold and cpa_change > CHANGE_MODERATE_PCT:
        add(
            "MEDIO",
            "Mantener presupuesto y revisar eficiencia antes de escalar.",
            f"CPA supera umbral y subió {cpa_change:.1f}%.",
            "presupuesto/eficiencia",
            confidence,
        )

    if roas < roas_threshold and roas_change < -CHANGE_MODERATE_PCT:
        add(
            "ALTO" if abs(roas_change) >= CHANGE_CRITICAL_PCT else "MEDIO",
            "Revisar distribución de presupuesto y calidad del tráfico. No aumentar inversión hasta recuperar ROAS.",
            f"ROAS bajo umbral y variación de {roas_change:.1f}%.",
            "rentabilidad",
            confidence,
        )

    if ctr < ctr_threshold or ctr_change < -CHANGE_MODERATE_PCT:
        add(
            "MEDIO" if abs(ctr_change) < CHANGE_CRITICAL_PCT else "ALTO",
            "Revisar títulos, descripciones, imágenes, fatiga creativa y relevancia del anuncio.",
            f"CTR actual {ctr:.2f}% y cambio {ctr_change:.1f}%.",
            "creatividad",
            confidence,
        )

    if clicks_change > CHANGE_MODERATE_PCT and conversions_change <= 0 and spend_change > 0:
        add(
            "ALTO" if clicks_change >= CHANGE_CRITICAL_PCT else "MEDIO",
            "Controlar keywords, términos de búsqueda, negativas, concordancias y calidad del tráfico.",
            f"Clics suben {clicks_change:.1f}%, gasto sube {spend_change:.1f}%, pero conversiones no acompañan.",
            "keywords/tráfico",
            confidence,
        )

    if cvr_change < -CHANGE_MODERATE_PCT and ctr_change >= -CHANGE_NOISE_PCT:
        add(
            "MEDIO",
            "Revisar landing, oferta, velocidad, formularios y evento de conversión.",
            f"CVR cae {cvr_change:.1f}% mientras el CTR no cae de forma equivalente.",
            "funnel/landing",
            confidence,
        )

    if (
        cpa_change < -CHANGE_MODERATE_PCT
        and roas_change > CHANGE_MODERATE_PCT
        and conversions_change > 0
        and cvr_change >= -CHANGE_NOISE_PCT
    ):
        add(
            "OPORTUNIDAD",
            f"Considerar aumento gradual de presupuesto de 5% a {MAX_SCALE_UP_PCT:.0f}% con revisión humana.",
            f"CPA baja {abs(cpa_change):.1f}%, ROAS sube {roas_change:.1f}% y conversiones acompañan.",
            "escalado",
            confidence,
        )

    if not rows:
        add(
            "NORMAL",
            "Mantener presupuesto y continuar monitoreo horario.",
            "No hay deterioro relevante por encima de los umbrales comerciales configurados.",
            "monitoreo",
            "BAJA" if history_len < 3 else "MEDIA",
            human_review=False,
        )

    return pd.DataFrame(rows)


def build_creative_assets(campaign_df: pd.DataFrame) -> pd.DataFrame:
    if campaign_df.empty:
        return pd.DataFrame()

    df = campaign_df.copy()

    expected_cols = [
        "ad_title",
        "ad_description",
        "image_asset_url",
        "image_file",
        "creative_group",
    ]

    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""

    def creative_recommendation(row):
        ctr = safe_float(row.get("ctr"))
        cvr = safe_float(row.get("cvr"))
        cpa = safe_float(row.get("cpa"))
        roas = safe_float(row.get("roas"))

        if ctr < 1.0:
            return "Revisar títulos, descripciones e imágenes. Posible baja relevancia o fatiga creativa."

        if ctr >= 1.0 and cvr < 1.0:
            return "El anuncio consigue clics, pero la conversión es débil. Revisar intención, landing u oferta."

        if cpa > 25 and roas < 1.5:
            return "Creatividad y tráfico no están generando eficiencia suficiente. Revisar ángulo, promesa y segmentación."

        return "Mantener creatividad bajo monitoreo."

    df["creative_recommendation"] = df.apply(creative_recommendation, axis=1)

    wanted_cols = [
        "summary_timestamp",
        "run_id",
        "client_number",
        "platform",
        "account_name",
        "campaign_name",
        "creative_group",
        "ad_title",
        "ad_description",
        "image_asset_url",
        "image_file",
        "spend",
        "impressions",
        "clicks",
        "ctr",
        "conversions",
        "cvr",
        "cpa",
        "roas",
        "creative_recommendation",
    ]

    for col in wanted_cols:
        if col not in df.columns:
            df[col] = ""

    return df[wanted_cols]


def normalize_df_for_fingerprint(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []

    normalized = df.copy()
    normalized = normalized.drop(columns=["timestamp_dt", "summary_timestamp_dt"], errors="ignore")
    normalized = normalized.sort_index(axis=1)

    for col in normalized.columns:
        normalized[col] = normalized[col].astype(str).fillna("")

    return normalized.to_dict(orient="records")


def build_fingerprint(
    hourly_df: pd.DataFrame,
    campaign_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    recommendations_df: pd.DataFrame,
    creative_df: pd.DataFrame,
) -> str:
    payload = {
        "hourly_summary": normalize_df_for_fingerprint(hourly_df),
        "campaign_metrics": normalize_df_for_fingerprint(campaign_df),
        "metric_comparison": normalize_df_for_fingerprint(comparison_df),
        "recommendations": normalize_df_for_fingerprint(recommendations_df),
        "creative_assets": normalize_df_for_fingerprint(creative_df),
    }

    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def load_previous_fingerprint() -> str:
    data = load_json_safely(FINGERPRINT_FILE)
    return str(data.get("fingerprint", ""))


def save_fingerprint(fingerprint: str, unique_runs: int):
    payload = {
        "fingerprint": fingerprint,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "unique_runs": unique_runs,
    }
    FINGERPRINT_FILE.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def autosize_and_style(path: Path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            width = min(max(max_length + 2, 10), 45)
            ws.column_dimensions[column_letter].width = width

    if "metric_comparison" in wb.sheetnames:
        ws = wb["metric_comparison"]
        # Columna I = change_pct en la versión actual.
        ws.conditional_formatting.add(
            "I2:I10000",
            CellIsRule(
                operator="greaterThan",
                formula=["10"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        ws.conditional_formatting.add(
            "I2:I10000",
            CellIsRule(
                operator="lessThan",
                formula=["-10"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    wb.save(path)


def write_workbook(
    path: Path,
    hourly_df: pd.DataFrame,
    campaign_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    recommendations_df: pd.DataFrame,
    creative_df: pd.DataFrame,
):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        hourly_df.to_excel(writer, sheet_name="hourly_summary", index=False)
        campaign_df.to_excel(writer, sheet_name="campaign_metrics", index=False)
        comparison_df.to_excel(writer, sheet_name="metric_comparison", index=False)
        recommendations_df.to_excel(writer, sheet_name="recommendations", index=False)
        creative_df.to_excel(writer, sheet_name="creative_assets", index=False)

    autosize_and_style(path)


def cleanup_old_pending_files():
    for pending in HISTORY_DIR.glob("HMA_Master_PENDING_*.xlsx"):
        try:
            pending.unlink()
        except PermissionError:
            pass


def save_master_safely(temp_file: Path, fingerprint: str, unique_runs: int):
    try:
        temp_file.replace(MASTER_FILE)
        save_fingerprint(fingerprint, unique_runs=unique_runs)
        cleanup_old_pending_files()
        print(f"HMA Master actualizado correctamente: {MASTER_FILE}")
        return
    except PermissionError:
        pending_name = f"HMA_Master_PENDING_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        pending_file = HISTORY_DIR / pending_name
        temp_file.replace(pending_file)

        print("HMA_Master.xlsx está abierto o bloqueado.")
        print("No se perdió la actualización.")
        print(f"Se guardó una copia alternativa: {pending_file}")
        print("Cuando cierres HMA_Master.xlsx, corré nuevamente el script para promover la actualización al archivo principal.")
        return


def main():
    runs = find_artifact_runs()

    hourly_df = build_hourly_summary(runs)
    campaign_df = build_campaign_metrics(runs)
    comparison_df = build_metric_comparison(hourly_df)
    recommendations_df = build_recommendations(hourly_df, comparison_df)
    creative_df = build_creative_assets(campaign_df)

    unique_runs = len(hourly_df) if not hourly_df.empty else 0

    fingerprint = build_fingerprint(
        hourly_df=hourly_df,
        campaign_df=campaign_df,
        comparison_df=comparison_df,
        recommendations_df=recommendations_df,
        creative_df=creative_df,
    )

    previous_fingerprint = load_previous_fingerprint()

    if MASTER_FILE.exists() and fingerprint == previous_fingerprint:
        print("No hay reportes horarios nuevos ni cambios analíticos.")
        print("No se actualiza HMA_Master.xlsx y no se crea PENDING.")
        print(f"Registros horarios únicos actuales: {unique_runs}")
        return

    temp_file = HISTORY_DIR / f"_HMA_Master_BUILDING_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    write_workbook(
        path=temp_file,
        hourly_df=hourly_df,
        campaign_df=campaign_df,
        comparison_df=comparison_df,
        recommendations_df=recommendations_df,
        creative_df=creative_df,
    )

    save_master_safely(temp_file, fingerprint=fingerprint, unique_runs=unique_runs)

    print(f"Artifacts únicos procesados: {len(runs)}")
    print(f"Registros horarios únicos en Excel: {unique_runs}")
    print(f"Filas hourly_summary: {len(hourly_df)}")
    print(f"Filas campaign_metrics: {len(campaign_df)}")
    print(f"Filas metric_comparison: {len(comparison_df)}")
    print(f"Filas recommendations: {len(recommendations_df)}")
    print(f"Filas creative_assets: {len(creative_df)}")


if __name__ == "__main__":
    main()
