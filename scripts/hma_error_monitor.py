from __future__ import annotations

import json
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
ERROR_DIR = BASE_DIR / "error"
DOWNLOADS_DIR = BASE_DIR / "downloads"
HISTORY_DIR = BASE_DIR / "historico"
MASTER_FILE = HISTORY_DIR / "HMA_Master.xlsx"

REPO = "UsuarioComun94/Google-ADS-analyzer-report-every-1-hour"
CLIENT_NUMBER = "CLIENTE-DEMO-0001"
REPORT_SUFFIX = "JPPQ"

# El reporte deberia estar disponible como maximo cerca del minuto 05.
# Desde minuto 06 en adelante, si falta algo, se registra error.
GRACE_MINUTE = 6


def now_local() -> datetime:
    return datetime.now()


def expected_context() -> dict[str, str]:
    now = now_local()
    hour_start = now.replace(minute=0, second=0, microsecond=0)

    date_slug = hour_start.strftime("%Y-%m-%d")
    hour_slug = hour_start.strftime("%H-00")
    artifact_name = f"Campania_{CLIENT_NUMBER}_{date_slug}_{hour_slug}_{REPORT_SUFFIX}"

    return {
        "date_slug": date_slug,
        "hour_slug": hour_slug,
        "hour_label": hour_start.strftime("%Y-%m-%d %H:00"),
        "timestamp_excel": hour_start.strftime("%Y-%m-%d %H:00:00"),
        "artifact_name": artifact_name,
    }


def ensure_error_dir() -> Path:
    ERROR_DIR.mkdir(parents=True, exist_ok=True)
    return ERROR_DIR


def write_error_file(error_type: str, ctx: dict[str, str], details: dict[str, Any]) -> Path:
    """
    Crea 1 TXT por cada deteccion de error.

    El nombre del archivo contiene solamente fecha y hora del chequeo:
    YYYY-MM-DD_HH-mm-ss.txt

    Si por alguna razon hay colision en el mismo segundo, agrega _01, _02, etc.
    """
    error_dir = ensure_error_dir()
    checked_at_dt = now_local()
    checked_at = checked_at_dt.strftime("%Y-%m-%d %H:%M:%S")
    base_name = checked_at_dt.strftime("%Y-%m-%d_%H-%M-%S")
    path = error_dir / f"{base_name}.txt"

    counter = 1
    while path.exists():
        path = error_dir / f"{base_name}_{counter:02d}.txt"
        counter += 1

    payload = {
        "checked_at": checked_at,
        "error_type": error_type,
        "expected_hour": ctx["hour_label"],
        "expected_artifact": ctx["artifact_name"],
        **details,
    }

    readable = [
        "HMA ERROR REPORT",
        "",
        f"Fecha de chequeo: {checked_at}",
        f"Hora esperada del reporte: {ctx['hour_label']}",
        f"Tipo de error: {error_type}",
        "",
        "Artifact esperado:",
        ctx["artifact_name"],
        "",
        "Detalle:",
        json.dumps(payload, ensure_ascii=False, indent=2),
        "",
    ]

    path.write_text("\n".join(readable), encoding="utf-8")

    index_path = ERROR_DIR / "hma_error_index.log"
    with index_path.open("a", encoding="utf-8") as f:
        f.write(f"[{checked_at}] {error_type} | expected_hour={ctx['hour_label']} | file={path.name}\n")

    return path


def run_gh_api(path: str) -> dict[str, Any]:
    result = subprocess.run(
        ["gh", "api", path],
        cwd=str(BASE_DIR),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=45,
    )

    if result.returncode != 0:
        raise RuntimeError(
            "No se pudo consultar GitHub con gh api. "
            f"returncode={result.returncode}; stderr={result.stderr.strip()}; stdout={result.stdout.strip()}"
        )

    return json.loads(result.stdout)


def find_expected_artifact(ctx: dict[str, str]) -> dict[str, Any] | None:
    data = run_gh_api(f"repos/{REPO}/actions/artifacts?per_page=100")
    artifacts = data.get("artifacts", [])

    for artifact in artifacts:
        if artifact.get("expired") is True:
            continue

        if artifact.get("name") == ctx["artifact_name"]:
            return artifact

    return None


def local_run_folder_exists(run_id: str) -> bool:
    if not run_id or not DOWNLOADS_DIR.exists():
        return False

    pattern = f"*_run-{run_id}"
    return any(DOWNLOADS_DIR.rglob(pattern))


def excel_has_hour(ctx: dict[str, str]) -> bool:
    if not MASTER_FILE.exists():
        return False

    try:
        wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    except Exception:
        return False

    if "hourly_summary" not in wb.sheetnames:
        return False

    ws = wb["hourly_summary"]

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    if "timestamp" not in headers:
        return False

    timestamp_col = headers.index("timestamp") + 1
    expected = ctx["timestamp_excel"]

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=timestamp_col).value

        if value is None:
            continue

        value_text = str(value).strip()

        if value_text.startswith(expected):
            return True

        # Tolerancia si Excel muestra sin segundos.
        if value_text.startswith(expected[:-3]):
            return True

    return False


def main() -> int:
    now = now_local()

    if now.minute < GRACE_MINUTE:
        print(
            f"Monitor HMA: minuto {now.minute}. "
            "Todavia dentro del margen operativo. No se registra error."
        )
        return 0

    ctx = expected_context()

    try:
        artifact = find_expected_artifact(ctx)
    except Exception as exc:
        path = write_error_file(
            "MONITOR_GITHUB_QUERY_FAILED",
            ctx,
            {
                "meaning": "El monitor local no pudo consultar GitHub para verificar si existe el artifact esperado.",
                "probable_cause": "Problema de GitHub CLI, autenticacion, internet, permisos o GitHub API.",
                "operational_impact": "No se puede confirmar si el reporte fue generado.",
                "exception": str(exc),
            },
        )
        print(f"ERROR registrado: {path}")
        return 1

    if artifact is None:
        path = write_error_file(
            "GITHUB_ARTIFACT_NOT_GENERATED",
            ctx,
            {
                "meaning": "No existe el artifact HMA esperado para la hora actual despues del margen operativo.",
                "expected_by": f"{ctx['hour_label'][:13]}:05",
                "probable_cause": "GitHub Actions no ejecuto a tiempo, el workflow fallo, el gate bloqueo, o el artifact no se subio.",
                "operational_impact": "El reporte no esta disponible para lectura operativa de la hora actual.",
            },
        )
        print(f"ERROR registrado: {path}")
        return 2

    workflow_run = artifact.get("workflow_run") or {}
    run_id = str(workflow_run.get("id", "")).strip()

    if not local_run_folder_exists(run_id):
        path = write_error_file(
            "ARTIFACT_EXISTS_BUT_NOT_DOWNLOADED",
            ctx,
            {
                "meaning": "GitHub genero el artifact esperado, pero no existe carpeta local descargada para ese run.",
                "run_id": run_id,
                "artifact_id": artifact.get("id", ""),
                "artifact_created_at": artifact.get("created_at", ""),
                "probable_cause": "La tarea local de descarga no corrio, fallo gh run download, hubo problema de red, o la PC estaba apagada/suspendida.",
                "operational_impact": "El reporte existe en GitHub, pero todavia no esta disponible localmente.",
            },
        )
        print(f"ERROR registrado: {path}")
        return 3

    if not excel_has_hour(ctx):
        path = write_error_file(
            "DOWNLOADED_BUT_EXCEL_NOT_UPDATED",
            ctx,
            {
                "meaning": "El artifact parece estar descargado localmente, pero HMA_Master.xlsx no contiene la fila de esa hora.",
                "run_id": run_id,
                "artifact_id": artifact.get("id", ""),
                "probable_cause": "Excel estaba abierto, update_hma_master.py fallo, hubo PENDING no promovido, o el archivo descargado no tenia estructura valida.",
                "operational_impact": "El reporte fue generado/descargado, pero el historico operativo no quedo actualizado.",
            },
        )
        print(f"ERROR registrado: {path}")
        return 4

    print(f"Monitor HMA OK: artifact, descarga local y Excel presentes para {ctx['hour_label']}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
