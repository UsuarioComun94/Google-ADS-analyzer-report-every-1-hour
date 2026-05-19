from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MASTER = Path(r"D:\Proyectos\hma-system\historico\HMA_Master.xlsx")
SRC = "metric_crosses"
DST = "recommendations"
MAX_PER_HOUR = 4

HEADERS = [
    "fecha_hora",
    "cliente",
    "resumen_horario",
    "prioridad",
    "tipo",
    "severidad",
    "area",
    "patron_detectado",
    "accion_recomendada",
    "accion_bloqueada",
    "motivo",
    "confianza",
    "requiere_revision_humana",
    "calidad_comparacion",
    "brecha_comparacion_horas",
    "estado_muestra",
    "metricas_fuente",
]

def clean(v):
    return "" if v is None else str(v).strip()

def key(v):
    return clean(v).lower()

def hmap(ws):
    return {key(c.value): c.column for c in ws[1] if c.value}

def get(ws, h, r, *names):
    for n in names:
        c = h.get(n)
        if c:
            return ws.cell(r, c).value
    return None

def as_float(v):
    try:
        if v is None or clean(v) == "":
            return None
        return float(str(v).replace(",", "."))
    except Exception:
        return None

def dt_sort(v):
    t = clean(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(t[:19], fmt)
        except Exception:
            pass
    return datetime.min

def tipo_from(sev):
    s = key(sev)
    if s in ("crítico", "critico", "critical", "alto", "alta", "high"):
        return "alerta"
    if s in ("oportunidad", "opportunity"):
        return "oportunidad"
    if s in ("advertencia", "warning"):
        return "advertencia"
    return "informativo"

def sev_rank(sev):
    s = key(sev)
    if s in ("crítico", "critico", "critical"):
        return 1
    if s in ("alto", "alta", "high"):
        return 2
    if s in ("oportunidad", "opportunity"):
        return 3
    if s in ("advertencia", "warning"):
        return 4
    return 5

def prioridad(sev, calidad, gap):
    s = key(sev)
    q = key(calidad)
    degraded = (gap is not None and gap > 1.5) or "degrad" in q or "missing" in q or "faltante" in q

    if degraded:
        return "P2 - revisar con cautela"
    if s in ("crítico", "critico", "critical"):
        return "P1 - acción inmediata"
    if s in ("alto", "alta", "high"):
        return "P2 - revisar hoy"
    if s in ("oportunidad", "opportunity"):
        return "P2 - oportunidad controlada"
    if s in ("advertencia", "warning"):
        return "P3 - monitorear"
    return "P4 - informativo"

def area(pattern, diag, metrics):
    t = f"{pattern} {diag} {metrics}".lower()
    if any(x in t for x in ("roas", "cpa", "revenue", "ingresos", "rentabilidad")):
        return "rentabilidad"
    if any(x in t for x in ("cvr", "conversion", "landing", "form", "oferta")):
        return "conversión"
    if any(x in t for x in ("ctr", "cpc", "click", "creativ", "keyword", "audiencia")):
        return "tráfico/creatividad"
    if any(x in t for x in ("brecha", "gap", "muestra", "sample", "missing")):
        return "calidad_datos"
    return "performance"

def summary(rows):
    severities = Counter(key(r["severidad"]) for r in rows)
    patterns = [r["patron_detectado"] for r in rows[:3]]

    crit = severities.get("crítico", 0) + severities.get("critico", 0)
    alto = severities.get("alto", 0) + severities.get("alta", 0)
    opp = severities.get("oportunidad", 0)
    warn = severities.get("advertencia", 0)

    parts = []
    if crit:
        parts.append(f"{crit} alerta(s) crítica(s)")
    if alto:
        parts.append(f"{alto} alerta(s) alta(s)")
    if opp:
        parts.append(f"{opp} oportunidad(es)")
    if warn:
        parts.append(f"{warn} advertencia(s)")

    base = "Hora con " + ", ".join(parts) if parts else "Hora sin señales fuertes"
    if patterns:
        base += ": " + "; ".join(dict.fromkeys(patterns)) + "."

    if crit:
        base += " Acción central: no escalar y auditar rentabilidad/tracking antes de mover presupuesto."
    elif opp:
        base += " Acción central: evaluar escala controlada con validación de estabilidad."
    elif warn:
        base += " Acción central: revisar calidad de datos antes de decidir."
    else:
        base += " Acción central: monitorear."

    return base

def style(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E8EEF7")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)

    widths = [22, 20, 80, 24, 16, 16, 22, 36, 70, 62, 82, 16, 24, 28, 22, 22, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    colors = {
        "crítico": ("FEE2E2", "991B1B"),
        "critico": ("FEE2E2", "991B1B"),
        "alto": ("FEF3C7", "92400E"),
        "alta": ("FEF3C7", "92400E"),
        "oportunidad": ("DCFCE7", "166534"),
        "advertencia": ("E0F2FE", "075985"),
    }

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 62
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name="Calibri", size=10, color="111827")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin, right=thin)

        sev = key(ws.cell(r, 6).value)
        if sev in colors:
            fill, font = colors[sev]
            for c in (5, 6):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
                ws.cell(r, c).font = Font(name="Calibri", size=10, bold=True, color=font)
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.sheet_properties.tabColor = "6A1B9A"

wb = load_workbook(MASTER)

if SRC not in wb.sheetnames:
    raise SystemExit("No existe metric_crosses. Frenar.")

src = wb[SRC]
h = hmap(src)

groups = defaultdict(list)

for r in range(2, src.max_row + 1):
    fecha = clean(get(src, h, r, "fecha_hora", "timestamp"))
    cliente = clean(get(src, h, r, "cliente", "client_number"))
    patron = clean(get(src, h, r, "patron", "patrón", "pattern"))

    if not fecha or not patron:
        continue

    sev = clean(get(src, h, r, "severidad", "severity"))
    conf = clean(get(src, h, r, "confianza", "confidence")) or "media"
    muestra = clean(get(src, h, r, "estado_muestra", "sample_status"))
    gap = as_float(get(src, h, r, "brecha_comparacion_horas", "comparison_gap_hours"))
    calidad = clean(get(src, h, r, "calidad_comparacion", "comparison_quality"))
    diag = clean(get(src, h, r, "diagnostico_probable", "diagnóstico_probable", "likely_bottleneck"))
    evidencia = clean(get(src, h, r, "evidencia", "evidence"))
    accion = clean(get(src, h, r, "accion_recomendada", "acción_recomendada", "recommended_action"))
    bloqueo = clean(get(src, h, r, "accion_bloqueada", "acción_bloqueada", "blocked_action"))
    metricas = clean(get(src, h, r, "metricas_fuente", "métricas_fuente", "source_metrics"))

    motivo = evidencia
    if diag:
        motivo += f" Diagnóstico probable: {diag}."
    if gap is not None and gap > 1.5:
        motivo += f" Comparación degradada: brecha de {gap:g} horas."

    row = {
        "fecha_hora": fecha,
        "cliente": cliente,
        "prioridad": prioridad(sev, calidad, gap),
        "tipo": tipo_from(sev),
        "severidad": sev or "informativo",
        "area": area(patron, diag, metricas),
        "patron_detectado": patron,
        "accion_recomendada": accion or "Revisar el cruce detectado y validar datos antes de ejecutar cambios.",
        "accion_bloqueada": bloqueo or "No ejecutar cambios agresivos sin validación.",
        "motivo": motivo.strip(),
        "confianza": conf,
        "requiere_revision_humana": "VERDADERO" if sev_rank(sev) <= 2 or (gap is not None and gap > 1.5) else "FALSO",
        "calidad_comparacion": calidad or "no disponible",
        "brecha_comparacion_horas": gap if gap is not None else "no disponible",
        "estado_muestra": muestra or "no disponible",
        "metricas_fuente": metricas or "no disponible",
    }

    groups[(fecha, cliente)].append(row)

if DST in wb.sheetnames:
    ws = wb[DST]
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet(DST)

ws.append(HEADERS)

total = 0

for (fecha, cliente), rows in sorted(groups.items(), key=lambda x: (dt_sort(x[0][0]), x[0][1])):
    rows = sorted(rows, key=lambda r: (sev_rank(r["severidad"]), r["patron_detectado"]))

    selected = []
    seen_patterns = set()

    for row in rows:
        if row["patron_detectado"] in seen_patterns:
            continue
        selected.append(row)
        seen_patterns.add(row["patron_detectado"])
        if len(selected) >= MAX_PER_HOUR:
            break

    resumen = summary(selected)

    for row in selected:
        ws.append([
            row["fecha_hora"],
            row["cliente"],
            resumen,
            row["prioridad"],
            row["tipo"],
            row["severidad"],
            row["area"],
            row["patron_detectado"],
            row["accion_recomendada"],
            row["accion_bloqueada"],
            row["motivo"],
            row["confianza"],
            row["requiere_revision_humana"],
            row["calidad_comparacion"],
            row["brecha_comparacion_horas"],
            row["estado_muestra"],
            row["metricas_fuente"],
        ])
        total += 1

style(ws)
wb.save(MASTER)

print(f"metric_crosses leídos: {src.max_row - 1}")
print(f"horas procesadas: {len(groups)}")
print(f"recommendations generadas: {total}")
print("recommendations reconstruida por fecha_hora + resumen_horario.")
