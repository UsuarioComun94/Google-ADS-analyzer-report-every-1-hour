from pathlib import Path
from datetime import datetime
import argparse
import json
import math
import shutil
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_MASTER = BASE_DIR / "historico" / "HMA_Master.xlsx"

SHEET_ORDER = [
    "hourly_summary",
    "campaign_metrics",
    "metric_comparison",
    "recommendations",
    "creative_assets",
    "metric_crosses",
]

HEADERS = {
    "hourly_summary": [
        "numero_cliente",
        "cliente",
        "fecha_hora",
        "time_zone",
        "health_status",
        "plataforma",
        "spend",
        "impressions",
        "clicks",
        "conversions",
        "ctr",
        "cvr",
        "cpa",
        "conversion_value",
        "roas",
    ],
    "campaign_metrics": [
        "numero_cliente",
        "cliente",
        "fecha_hora",
        "plataforma",
        "campaign_id",
        "campaign_name",
        "spend",
        "impressions",
        "clicks",
        "conversions",
        "ctr",
        "cvr",
        "cpa",
        "conversion_value",
        "roas",
        "source_file",
    ],
    "metric_comparison": [
        "fecha_hora",
        "cliente",
        "plataforma",
        "campaign_id",
        "campaign_name",
        "metrica",
        "valor_actual",
        "valor_anterior",
        "cambio_absoluto",
        "cambio_pct",
        "estado",
        "interpretacion",
    ],
    "recommendations": [
        "fecha_hora",
        "cliente",
        "resumen_horario",
        "prioridad",
        "tipo",
        "severidad",
        "area",
        "patron_detectado",
        "accion_recomendada",
        "accion_bloqueada",
        "motivo",
        "confianza",
    ],
    "creative_assets": [
        "fecha_hora",
        "cliente",
        "plataforma",
        "campaign_id",
        "campaign_name",
        "estado_creativo",
        "senal_detectada",
        "accion_recomendada",
        "motivo",
        "confianza",
    ],
    "metric_crosses": [
        "id_cruce",
        "fecha_hora",
        "cliente",
        "plataforma",
        "campaign_id",
        "campaign_name",
        "patron_detectado",
        "metricas_cruzadas",
        "interpretacion",
        "accion_recomendada",
        "accion_bloqueada",
        "motivo",
        "confianza",
    ],
}


def safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        if isinstance(value, float) and math.isnan(value):
            return default
        if str(value).strip() == "":
            return default
        return float(value)
    except Exception:
        return default


def safe_div(a, b):
    a = safe_float(a)
    b = safe_float(b)
    if b == 0:
        return 0.0
    return a / b


def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def normalize_hour(value):
    try:
        if value is None or str(value).strip() == "":
            return ""
        return f"{int(float(value)):02d}:00"
    except Exception:
        return safe_str(value)


def build_fecha_hora(date_value, hour_value=""):
    date_text = safe_str(date_value).strip()
    hour_text = normalize_hour(hour_value)

    if not date_text:
        return ""

    if hour_text:
        return f"{date_text} {hour_text}"

    return date_text


def read_csv_folder(folder: Path) -> pd.DataFrame:
    files = sorted(folder.glob("*.csv"))
    frames = []

    for f in files:
        try:
            df = pd.read_csv(f)
            df["source_file"] = f.name
            frames.append(df)
        except Exception as exc:
            print(f"ERROR_READING_CSV file={f} error={exc}")

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def normalize_google(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["extracted_at"] = df.get("extracted_at", "")
    out["plataforma"] = "google_ads"
    out["date"] = df.get("date", "")
    out["hour"] = df.get("hour", "")
    out["fecha_hora"] = [
        build_fecha_hora(d, h)
        for d, h in zip(out["date"], out["hour"])
    ]
    out["campaign_id"] = df.get("campaign_id", "")
    out["campaign_name"] = df.get("campaign_name", "")
    out["spend"] = df.get("spend", 0).apply(safe_float)
    out["impressions"] = df.get("impressions", 0).apply(safe_float)
    out["clicks"] = df.get("clicks", 0).apply(safe_float)
    out["conversions"] = df.get("conversions", 0).apply(safe_float)
    out["conversion_value"] = df.get("conversion_value", 0).apply(safe_float)
    out["source_file"] = df.get("source_file", "")

    return out


def normalize_meta(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["extracted_at"] = df.get("extracted_at", "")
    out["plataforma"] = "meta_ads"
    out["date"] = df.get("date_start", "")
    out["hour"] = ""
    out["fecha_hora"] = [build_fecha_hora(d, "") for d in out["date"]]
    out["campaign_id"] = df.get("campaign_id", "")
    out["campaign_name"] = df.get("campaign_name", "")
    out["spend"] = df.get("spend", 0).apply(safe_float)
    out["impressions"] = df.get("impressions", 0).apply(safe_float)
    out["clicks"] = df.get("clicks", 0).apply(safe_float)
    out["conversions"] = df.get("conversions", 0).apply(safe_float)
    out["conversion_value"] = df.get("conversion_value", 0).apply(safe_float)
    out["source_file"] = df.get("source_file", "")

    return out


def add_calculated_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df["ctr"] = df.apply(lambda r: safe_div(r["clicks"], r["impressions"]), axis=1)
    df["cvr"] = df.apply(lambda r: safe_div(r["conversions"], r["clicks"]), axis=1)
    df["cpa"] = df.apply(lambda r: safe_div(r["spend"], r["conversions"]), axis=1)
    df["roas"] = df.apply(lambda r: safe_div(r["conversion_value"], r["spend"]), axis=1)

    return df


def build_campaign_metrics(client_id, client_name, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=HEADERS["campaign_metrics"])

    rows = []

    for _, r in df.iterrows():
        rows.append({
            "numero_cliente": client_id,
            "cliente": client_name,
            "fecha_hora": r.get("fecha_hora", ""),
            "plataforma": r.get("plataforma", ""),
            "campaign_id": r.get("campaign_id", ""),
            "campaign_name": r.get("campaign_name", ""),
            "spend": safe_float(r.get("spend")),
            "impressions": safe_float(r.get("impressions")),
            "clicks": safe_float(r.get("clicks")),
            "conversions": safe_float(r.get("conversions")),
            "ctr": safe_float(r.get("ctr")),
            "cvr": safe_float(r.get("cvr")),
            "cpa": safe_float(r.get("cpa")),
            "conversion_value": safe_float(r.get("conversion_value")),
            "roas": safe_float(r.get("roas")),
            "source_file": r.get("source_file", ""),
        })

    return pd.DataFrame(rows, columns=HEADERS["campaign_metrics"])


def build_hourly_summary(client_id, client_name, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=HEADERS["hourly_summary"])

    grouped = (
        df.groupby(["fecha_hora", "plataforma"], dropna=False)
        .agg({
            "spend": "sum",
            "impressions": "sum",
            "clicks": "sum",
            "conversions": "sum",
            "conversion_value": "sum",
        })
        .reset_index()
    )

    grouped["ctr"] = grouped.apply(lambda r: safe_div(r["clicks"], r["impressions"]), axis=1)
    grouped["cvr"] = grouped.apply(lambda r: safe_div(r["conversions"], r["clicks"]), axis=1)
    grouped["cpa"] = grouped.apply(lambda r: safe_div(r["spend"], r["conversions"]), axis=1)
    grouped["roas"] = grouped.apply(lambda r: safe_div(r["conversion_value"], r["spend"]), axis=1)

    rows = []

    for _, r in grouped.iterrows():
        spend = safe_float(r.get("spend"))
        clicks = safe_float(r.get("clicks"))
        conversions = safe_float(r.get("conversions"))

        if spend > 0 and clicks == 0:
            status = "Riesgo"
        elif clicks > 0 and conversions == 0:
            status = "Atencion"
        else:
            status = "OK"

        rows.append({
            "numero_cliente": client_id,
            "cliente": client_name,
            "fecha_hora": r.get("fecha_hora", ""),
            "time_zone": "local",
            "health_status": status,
            "plataforma": r.get("plataforma", ""),
            "spend": spend,
            "impressions": safe_float(r.get("impressions")),
            "clicks": clicks,
            "conversions": conversions,
            "ctr": safe_float(r.get("ctr")),
            "cvr": safe_float(r.get("cvr")),
            "cpa": safe_float(r.get("cpa")),
            "conversion_value": safe_float(r.get("conversion_value")),
            "roas": safe_float(r.get("roas")),
        })

    return pd.DataFrame(rows, columns=HEADERS["hourly_summary"])


def build_metric_comparison(client_name, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=HEADERS["metric_comparison"])

    df = df.sort_values(["plataforma", "campaign_id", "fecha_hora"]).copy()
    rows = []

    metrics = ["spend", "clicks", "conversions", "ctr", "cvr", "cpa", "roas"]

    for _, group in df.groupby(["plataforma", "campaign_id"], dropna=False):
        group = group.sort_values("fecha_hora")

        previous = None

        for _, current in group.iterrows():
            if previous is None:
                previous = current
                continue

            for metric in metrics:
                actual = safe_float(current.get(metric))
                prev = safe_float(previous.get(metric))
                diff = actual - prev
                pct = safe_div(diff, prev) if prev != 0 else 0.0

                if abs(pct) >= 0.10:
                    estado = "Cambio critico"
                elif abs(pct) >= 0.05:
                    estado = "Cambio moderado"
                elif abs(pct) >= 0.02:
                    estado = "Cambio leve"
                else:
                    estado = "Sin cambio relevante"

                rows.append({
                    "fecha_hora": current.get("fecha_hora", ""),
                    "cliente": client_name,
                    "plataforma": current.get("plataforma", ""),
                    "campaign_id": current.get("campaign_id", ""),
                    "campaign_name": current.get("campaign_name", ""),
                    "metrica": metric,
                    "valor_actual": actual,
                    "valor_anterior": prev,
                    "cambio_absoluto": diff,
                    "cambio_pct": pct,
                    "estado": estado,
                    "interpretacion": f"{metric}: actual {actual:.4f} vs anterior {prev:.4f}.",
                })

            previous = current

    return pd.DataFrame(rows, columns=HEADERS["metric_comparison"])


def build_recommendations(client_name, hourly_df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    if hourly_df.empty:
        return pd.DataFrame([{
            "fecha_hora": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "cliente": client_name,
            "resumen_horario": "No hay datos exportados para analizar.",
            "prioridad": "INFO",
            "tipo": "datos",
            "severidad": "baja",
            "area": "tracking",
            "patron_detectado": "Sin datos",
            "accion_recomendada": "Conectar Google Ads o Meta Ads y ejecutar una exportacion.",
            "accion_bloqueada": "",
            "motivo": "No existen CSV en raw_exports para este cliente.",
            "confianza": 1.0,
        }], columns=HEADERS["recommendations"])

    for _, r in hourly_df.iterrows():
        fecha_hora = r.get("fecha_hora", "")
        plataforma = r.get("plataforma", "")
        spend = safe_float(r.get("spend"))
        clicks = safe_float(r.get("clicks"))
        conversions = safe_float(r.get("conversions"))
        ctr = safe_float(r.get("ctr"))
        cvr = safe_float(r.get("cvr"))
        cpa = safe_float(r.get("cpa"))
        roas = safe_float(r.get("roas"))

        resumen = (
            f"{plataforma} | Spend ${spend:.2f}, clicks {clicks:.0f}, "
            f"conversiones {conversions:.2f}, CTR {ctr:.2%}, CVR {cvr:.2%}, "
            f"CPA ${cpa:.2f}, ROAS {roas:.2f}."
        )

        if spend > 0 and clicks == 0:
            rows.append({
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "resumen_horario": resumen,
                "prioridad": "ALTA",
                "tipo": "alerta",
                "severidad": "alta",
                "area": "delivery/tracking",
                "patron_detectado": "Spend sin clicks",
                "accion_recomendada": "Revisar estado de campañas, segmentacion, aprobaciones y tracking.",
                "accion_bloqueada": "Escalar presupuesto",
                "motivo": "Hay inversion registrada sin respuesta de usuarios.",
                "confianza": 0.90,
            })

        if clicks > 0 and conversions == 0:
            rows.append({
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "resumen_horario": resumen,
                "prioridad": "MEDIA",
                "tipo": "diagnostico",
                "severidad": "media",
                "area": "conversion",
                "patron_detectado": "Clicks sin conversiones",
                "accion_recomendada": "Revisar landing, oferta, formulario, pixel/conversion tracking y calidad del trafico.",
                "accion_bloqueada": "Subir presupuesto",
                "motivo": "Existe trafico pero no se registran conversiones.",
                "confianza": 0.85,
            })

        if ctr > 0.05 and cvr < 0.01 and clicks >= 50:
            rows.append({
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "resumen_horario": resumen,
                "prioridad": "MEDIA",
                "tipo": "cruce_metricas",
                "severidad": "media",
                "area": "landing/oferta",
                "patron_detectado": "CTR alto + CVR bajo",
                "accion_recomendada": "Auditar landing, promesa del anuncio, formulario y friccion de conversion.",
                "accion_bloqueada": "Escalar presupuesto",
                "motivo": "El anuncio atrae clicks, pero la experiencia posterior no convierte.",
                "confianza": 0.82,
            })

    if not rows:
        rows.append({
            "fecha_hora": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "cliente": client_name,
            "resumen_horario": "No se detectaron alertas criticas con los datos disponibles.",
            "prioridad": "INFO",
            "tipo": "monitoreo",
            "severidad": "baja",
            "area": "general",
            "patron_detectado": "Sin alertas criticas",
            "accion_recomendada": "Mantener monitoreo y validar consistencia de datos.",
            "accion_bloqueada": "",
            "motivo": "Las reglas actuales no detectaron patrones criticos.",
            "confianza": 0.75,
        })

    return pd.DataFrame(rows, columns=HEADERS["recommendations"])


def build_creative_assets(client_name, campaign_df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    if campaign_df.empty:
        return pd.DataFrame(columns=HEADERS["creative_assets"])

    for _, r in campaign_df.iterrows():
        ctr = safe_float(r.get("ctr"))
        clicks = safe_float(r.get("clicks"))
        conversions = safe_float(r.get("conversions"))

        if clicks >= 50 and conversions == 0:
            estado = "Riesgo"
            senal = "Trafico sin conversion"
            accion = "Revisar mensaje, creatividad, landing y evento de conversion."
            motivo = "Hay volumen de clicks sin conversiones."
            confianza = 0.80
        elif ctr < 0.01 and safe_float(r.get("impressions")) >= 1000:
            estado = "Atencion"
            senal = "CTR bajo"
            accion = "Probar nuevos angulos creativos, anuncios o segmentacion."
            motivo = "La pieza no esta capturando suficiente interes."
            confianza = 0.75
        else:
            estado = "OK"
            senal = "Sin senal critica"
            accion = "Mantener monitoreo."
            motivo = "No se detecta fatiga o bloqueo con las reglas actuales."
            confianza = 0.60

        rows.append({
            "fecha_hora": r.get("fecha_hora", ""),
            "cliente": client_name,
            "plataforma": r.get("plataforma", ""),
            "campaign_id": r.get("campaign_id", ""),
            "campaign_name": r.get("campaign_name", ""),
            "estado_creativo": estado,
            "senal_detectada": senal,
            "accion_recomendada": accion,
            "motivo": motivo,
            "confianza": confianza,
        })

    return pd.DataFrame(rows, columns=HEADERS["creative_assets"])


def build_metric_crosses(client_name, campaign_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    counter = 1

    if campaign_df.empty:
        return pd.DataFrame(columns=HEADERS["metric_crosses"])

    for _, r in campaign_df.iterrows():
        fecha_hora = r.get("fecha_hora", "")
        plataforma = r.get("plataforma", "")
        campaign_id = r.get("campaign_id", "")
        campaign_name = r.get("campaign_name", "")

        spend = safe_float(r.get("spend"))
        clicks = safe_float(r.get("clicks"))
        conversions = safe_float(r.get("conversions"))
        ctr = safe_float(r.get("ctr"))
        cvr = safe_float(r.get("cvr"))
        cpa = safe_float(r.get("cpa"))
        roas = safe_float(r.get("roas"))

        if spend > 0 and clicks == 0:
            rows.append({
                "id_cruce": f"CR-{counter:05d}",
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "plataforma": plataforma,
                "campaign_id": campaign_id,
                "campaign_name": campaign_name,
                "patron_detectado": "Spend sin clicks",
                "metricas_cruzadas": "spend + clicks",
                "interpretacion": "Hay inversion registrada sin clicks. Puede ser problema de entrega, configuracion o tracking.",
                "accion_recomendada": "Revisar estado de campaña, segmentacion, anuncios y medicion.",
                "accion_bloqueada": "Escalar presupuesto",
                "motivo": f"Spend={spend:.2f}, clicks={clicks:.0f}",
                "confianza": 0.90,
            })
            counter += 1

        if clicks > 0 and conversions == 0:
            rows.append({
                "id_cruce": f"CR-{counter:05d}",
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "plataforma": plataforma,
                "campaign_id": campaign_id,
                "campaign_name": campaign_name,
                "patron_detectado": "Clicks sin conversiones",
                "metricas_cruzadas": "clicks + conversions + cpa",
                "interpretacion": "Existe trafico pero no se registran conversiones.",
                "accion_recomendada": "Auditar landing, oferta, formulario y conversion tracking.",
                "accion_bloqueada": "Subir presupuesto",
                "motivo": f"Clicks={clicks:.0f}, conversions={conversions:.2f}",
                "confianza": 0.85,
            })
            counter += 1

        if ctr > 0.05 and cvr < 0.01 and clicks >= 50:
            rows.append({
                "id_cruce": f"CR-{counter:05d}",
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "plataforma": plataforma,
                "campaign_id": campaign_id,
                "campaign_name": campaign_name,
                "patron_detectado": "CTR alto + CVR bajo",
                "metricas_cruzadas": "ctr + cvr + clicks",
                "interpretacion": "El anuncio genera interes, pero el proceso posterior no convierte.",
                "accion_recomendada": "Revisar consistencia anuncio-landing, promesa, formulario y friccion.",
                "accion_bloqueada": "Escalar presupuesto",
                "motivo": f"CTR={ctr:.2%}, CVR={cvr:.2%}, clicks={clicks:.0f}",
                "confianza": 0.82,
            })
            counter += 1

        if conversions > 0 and roas == 0:
            rows.append({
                "id_cruce": f"CR-{counter:05d}",
                "fecha_hora": fecha_hora,
                "cliente": client_name,
                "plataforma": plataforma,
                "campaign_id": campaign_id,
                "campaign_name": campaign_name,
                "patron_detectado": "Conversiones sin valor",
                "metricas_cruzadas": "conversions + conversion_value + roas",
                "interpretacion": "Hay conversiones pero no se registra valor. El ROAS queda incompleto.",
                "accion_recomendada": "Revisar value tracking, evento de compra/lead y configuracion de conversion value.",
                "accion_bloqueada": "",
                "motivo": f"Conversions={conversions:.2f}, ROAS={roas:.2f}, CPA={cpa:.2f}",
                "confianza": 0.78,
            })
            counter += 1

    return pd.DataFrame(rows, columns=HEADERS["metric_crosses"])


def load_client_config(client_dir: Path) -> dict:
    config_path = client_dir / "config" / "client_config.json"

    if not config_path.exists():
        return {}

    try:
        return json.loads(config_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}


def create_workbook_from_template(output: Path):
    output.parent.mkdir(parents=True, exist_ok=True)

    if TEMPLATE_MASTER.exists():
        shutil.copy2(TEMPLATE_MASTER, output)
        wb = load_workbook(output)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name in SHEET_ORDER:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        ws.sheet_state = "visible"

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        existing_headers = [cell.value for cell in ws[1]]
        if not existing_headers or all(v is None for v in existing_headers):
            for col_idx, header in enumerate(HEADERS[sheet_name], start=1):
                ws.cell(row=1, column=col_idx).value = header
        else:
            for col_idx, header in enumerate(HEADERS[sheet_name], start=1):
                ws.cell(row=1, column=col_idx).value = header

    for idx, sheet_name in enumerate(SHEET_ORDER):
        ws = wb[sheet_name]
        wb._sheets.remove(ws)
        wb._sheets.insert(idx, ws)

    return wb


def write_sheet(ws, df: pd.DataFrame):
    headers = [cell.value for cell in ws[1] if cell.value is not None]

    if not headers:
        return

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    if df.empty:
        return

    records = df.to_dict("records")

    for row_idx, record in enumerate(records, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = record.get(header, "")
            ws.cell(row=row_idx, column=col_idx).value = value


def apply_style(wb):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 10

            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                cell = ws.cell(row=row_idx, column=col_idx)

                if header in ["spend", "cpa", "conversion_value", "valor_actual", "valor_anterior", "cambio_absoluto"]:
                    cell.number_format = '$#,##0.00'
                elif header in ["ctr", "cvr", "roas", "cambio_pct", "confianza"]:
                    cell.number_format = '0.00%'
                elif header in ["impressions", "clicks", "conversions"]:
                    cell.number_format = '#,##0'

        if ws.title == "metric_crosses":
            ws.sheet_state = "visible"
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == "id_cruce":
                    ws.column_dimensions[get_column_letter(col_idx)].hidden = True

        if ws.title == "recommendations":
            priority_col = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == "prioridad":
                    priority_col = col_idx
                    break

            if priority_col and ws.max_row > 1:
                col_letter = get_column_letter(priority_col)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=['"ALTA"'], fill=PatternFill("solid", fgColor="FCA5A5"))
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=['"MEDIA"'], fill=PatternFill("solid", fgColor="FDE68A"))
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=['"INFO"'], fill=PatternFill("solid", fgColor="BFDBFE"))
                )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--client-dir", required=True)
    args = parser.parse_args()

    client_dir = Path(args.client_dir)
    config = load_client_config(client_dir)

    client_id = config.get("client_id", client_dir.name)
    client_name = config.get("client_name", client_dir.name)

    google_raw = read_csv_folder(client_dir / "raw_exports" / "google_ads")
    meta_raw = read_csv_folder(client_dir / "raw_exports" / "meta_ads")

    google_norm = normalize_google(google_raw)
    meta_norm = normalize_meta(meta_raw)

    if not google_norm.empty or not meta_norm.empty:
        normalized = pd.concat([google_norm, meta_norm], ignore_index=True)
    else:
        normalized = pd.DataFrame()

    normalized = add_calculated_columns(normalized)

    campaign_df = build_campaign_metrics(client_id, client_name, normalized)
    hourly_df = build_hourly_summary(client_id, client_name, normalized)
    comparison_df = build_metric_comparison(client_name, normalized)
    recommendations_df = build_recommendations(client_name, hourly_df)
    creative_df = build_creative_assets(client_name, campaign_df)
    crosses_df = build_metric_crosses(client_name, campaign_df)

    output = client_dir / "historico" / "HMA_Master.xlsx"

    wb = create_workbook_from_template(output)

    write_sheet(wb["hourly_summary"], hourly_df)
    write_sheet(wb["campaign_metrics"], campaign_df)
    write_sheet(wb["metric_comparison"], comparison_df)
    write_sheet(wb["recommendations"], recommendations_df)
    write_sheet(wb["creative_assets"], creative_df)
    write_sheet(wb["metric_crosses"], crosses_df)

    apply_style(wb)

    tmp = output.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(output)

    print("CLIENT_MASTER_OK")
    print(f"client={client_id} | {client_name}")
    print(f"file={output}")
    print(f"hourly_summary_rows={len(hourly_df)}")
    print(f"campaign_metrics_rows={len(campaign_df)}")
    print(f"metric_comparison_rows={len(comparison_df)}")
    print(f"recommendations_rows={len(recommendations_df)}")
    print(f"creative_assets_rows={len(creative_df)}")
    print(f"metric_crosses_rows={len(crosses_df)}")


if __name__ == "__main__":
    main()
