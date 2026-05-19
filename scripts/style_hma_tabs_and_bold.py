from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(r"D:\Proyectos\hma-system")
MASTER_PATH = BASE_DIR / "historico" / "HMA_Master.xlsx"
BACKUP_DIR = BASE_DIR / "historico" / "_visual_style_backups"

TAB_COLORS = {
    "hourly_summary": "2E7D32",
    "campaign_metrics": "1565C0",
    "metric_comparison": "EF6C00",
    "recommendations": "6A1B9A",
    "creative_assets": "455A64",
    "metric_crosses": "B71C1C",
}

HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
GRID = "D9E2F3"

HEALTH_GREEN = "DCFCE7"
HEALTH_YELLOW = "FEF3C7"
HEALTH_RED = "FEE2E2"
HEALTH_GRAY = "F1F5F9"

RELEVANCE_CRITICAL = "FEE2E2"
RELEVANCE_MODERATE = "FEF3C7"
RELEVANCE_LIGHT = "E0F2FE"
RELEVANCE_NORMAL = "F8FAFC"

MONEY_FILL = "EAF7EF"
METRIC_FILL = "EEF5FF"
TEXT_FILL = "FAFAFA"
SOFT_FILL = "F6F8FB"
WHITE_FILL = "FFFFFF"

HIDE_COLUMNS_BY_SHEET = {
    "hourly_summary": {
        "dedupe_key",
        "run_id",
        "run_key",
        "data_source",
        "requested_data_source",
    },
    "campaign_metrics": {
        "dedupe_key",
        "run_id",
        "run_key",
    },
    "metric_comparison": {
        "run_id",
        "compared_against_run_id",
    },
    "creative_assets": {
        "dedupe_key",
        "run_id",
        "run_key",
    },
}

MONEY_COLUMNS = {
    "spend",
    "revenue",
    "cpc",
    "cpa",
    "daily_budget_usd",
    "hourly_budget_usd",
    "current_value",
    "previous_value",
}

PERCENT_COLUMNS = {
    "ctr",
    "cvr",
    "change_pct",
}

ROAS_COLUMNS = {
    "roas",
}

INT_COLUMNS = {
    "impressions",
    "clicks",
    "conversions",
}

DATE_COLUMNS = {
    "timestamp",
    "compared_against_timestamp",
}

KEY_BOLD_COLUMNS = {
    "timestamp",
    "client_number",
    "campaign_name",
    "health_status",
    "metric",
    "relevance",
    "direction",
    "severity",
    "cross_id",
    "pattern",
}

TEXT_COLUMNS = {
    "campaign_name",
    "ad_title",
    "ad_description",
    "image_asset_url",
    "commercial_interpretation",
    "recommendation",
    "recommended_action",
    "blocked_actions",
    "evidence",
    "likely_bottleneck",
}


def norm(value) -> str:
    return "" if value is None else str(value).strip()


def lower(value) -> str:
    return norm(value).lower()


def headers(ws) -> dict[str, int]:
    result = {}
    for cell in ws[1]:
        key = lower(cell.value)
        if key:
            result[key] = cell.column
    return result


def backup() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dst = BACKUP_DIR / f"HMA_Master_pre_final_visual_style_{stamp}.xlsx"
    shutil.copy2(MASTER_PATH, dst)
    print(f"Backup creado: {dst}")


def restore_timestamp_header_if_needed(ws) -> None:
    if norm(ws["A1"].value):
        return

    sample = " ".join(
        norm(ws.cell(row=r, column=1).value)
        for r in range(2, min(ws.max_row, 8) + 1)
    )

    if "2026-" in sample or ":" in sample:
        ws["A1"] = "timestamp"
        print(f"Header timestamp restaurado en {ws.title}!A1")


def apply_tab_color(ws) -> None:
    color = TAB_COLORS.get(ws.title)
    if color:
        ws.sheet_properties.tabColor = color


def style_headers(ws) -> None:
    thin = Side(style="thin", color=GRID)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Calibri", size=10, bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)

    ws.row_dimensions[1].height = 30


def hide_technical_columns(ws) -> None:
    h = headers(ws)
    to_hide = HIDE_COLUMNS_BY_SHEET.get(ws.title, set())

    for name, col_idx in h.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter

        if name in to_hide:
            ws.column_dimensions[col_letter].hidden = True
            print(f"Oculta: {ws.title}!{col_letter} ({name})")
        else:
            # No forzar visibilidad global: solo ocultar columnas técnicas.
            # La visibilidad operativa se repara en repair_hma_layout_visibility.py.
            pass

    if lower(ws["A1"].value) == "timestamp":
        ws.column_dimensions["A"].hidden = False


def apply_body_style(ws) -> None:
    thin = Side(style="thin", color="E8EEF7")

    for col_idx in range(1, ws.max_column + 1):
        header = lower(ws.cell(row=1, column=col_idx).value)

        if header in MONEY_COLUMNS:
            fill = MONEY_FILL
        elif header in PERCENT_COLUMNS or header in ROAS_COLUMNS or header in INT_COLUMNS:
            fill = METRIC_FILL
        elif header in TEXT_COLUMNS or "description" in header or "interpretation" in header:
            fill = TEXT_FILL
        elif col_idx % 2 == 0:
            fill = SOFT_FILL
        else:
            fill = WHITE_FILL

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin, right=thin)
            cell.font = Font(name="Calibri", size=10, color="111827")
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 38


def apply_bold_columns(ws) -> None:
    h = headers(ws)

    for name, col_idx in h.items():
        if name not in KEY_BOLD_COLUMNS:
            continue

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10, bold=True, color="111827")


def apply_number_formats(ws) -> None:
    h = headers(ws)

    money_fmt = '"$" #,##0.00;[Red]-"$" #,##0.00'
    percent_fmt = '0.00"%"'
    roas_fmt = '0.00x'
    int_fmt = '#,##0'
    date_fmt = 'yyyy-mm-dd hh:mm'
    decimal_fmt = '#,##0.00'

    metric_col = h.get("metric")

    for name, col_idx in h.items():
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # En metric_comparison, current_value/previous_value dependen de la métrica de la fila.
            if ws.title == "metric_comparison" and name in {"current_value", "previous_value"} and metric_col:
                metric = lower(ws.cell(row=row_idx, column=metric_col).value)

                if metric in {"spend", "cpc", "cpa", "revenue"}:
                    cell.number_format = money_fmt
                elif metric in {"ctr", "cvr", "change_pct"}:
                    cell.number_format = percent_fmt
                elif metric == "roas":
                    cell.number_format = roas_fmt
                elif metric in {"impressions", "clicks", "conversions"}:
                    cell.number_format = int_fmt
                else:
                    cell.number_format = decimal_fmt

                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                continue

            if name in MONEY_COLUMNS:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            elif name in PERCENT_COLUMNS:
                cell.number_format = percent_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            elif name in ROAS_COLUMNS:
                cell.number_format = roas_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            elif name in INT_COLUMNS:
                cell.number_format = int_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            elif name in DATE_COLUMNS:
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)):
                cell.number_format = decimal_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def color_health_status(ws) -> None:
    h = headers(ws)
    col = h.get("health_status")

    if not col:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col)
        value = lower(cell.value)

        if value in {"normal", "ok", "healthy", "success"}:
            fill = HEALTH_GREEN
            font_color = "166534"
        elif value in {"warning", "warn", "medium", "degraded"}:
            fill = HEALTH_YELLOW
            font_color = "92400E"
        elif value in {"error", "critical", "critico", "crítico", "failed", "fail"}:
            fill = HEALTH_RED
            font_color = "991B1B"
        else:
            fill = HEALTH_GRAY
            font_color = "334155"

        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=10, bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def color_relevance_direction_severity(ws) -> None:
    h = headers(ws)

    relevance_col = h.get("relevance")
    direction_col = h.get("direction")
    severity_col = h.get("severity")

    if relevance_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=relevance_col)
            value = lower(cell.value)

            if "crítica" in value or "critica" in value:
                cell.fill = PatternFill("solid", fgColor=RELEVANCE_CRITICAL)
                cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")
            elif "moderada" in value or "moderado" in value:
                cell.fill = PatternFill("solid", fgColor=RELEVANCE_MODERATE)
                cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
            elif "leve" in value:
                cell.fill = PatternFill("solid", fgColor=RELEVANCE_LIGHT)
                cell.font = Font(name="Calibri", size=10, bold=True, color="075985")
            elif "ruido" in value:
                cell.fill = PatternFill("solid", fgColor=RELEVANCE_NORMAL)
                cell.font = Font(name="Calibri", size=10, color="475569")

    if direction_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=direction_col)
            value = lower(cell.value)

            if "subió" in value or "subio" in value:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
            elif "bajó" in value or "bajo" in value:
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
                cell.font = Font(name="Calibri", size=10, bold=True, color="1D4ED8")
            elif "estable" in value:
                cell.fill = PatternFill("solid", fgColor="E2E8F0")
                cell.font = Font(name="Calibri", size=10, bold=True, color="334155")

    if severity_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=severity_col)
            value = lower(cell.value)

            if "critical" in value or "crítico" in value or "critico" in value or "alto" in value:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
                cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")
            elif "medium" in value or "medio" in value or "moderado" in value:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
            elif "low" in value or "bajo" in value:
                cell.fill = PatternFill("solid", fgColor="E0F2FE")
                cell.font = Font(name="Calibri", size=10, bold=True, color="075985")


def set_widths(ws) -> None:
    h = headers(ws)

    widths = {
        "timestamp": 22,
        "client_number": 20,
        "timezone": 24,
        "health_status": 16,
        "platform": 15,
        "account_name": 20,
        "campaign_name": 34,
        "spend": 14,
        "daily_budget_usd": 16,
        "hourly_budget_usd": 16,
        "impressions": 14,
        "clicks": 12,
        "conversions": 14,
        "revenue": 14,
        "ctr": 12,
        "cpc": 12,
        "cvr": 12,
        "cpa": 12,
        "roas": 12,
        "creative_group": 22,
        "ad_title": 34,
        "ad_description": 48,
        "image_asset_url": 42,
        "metric": 15,
        "current_value": 15,
        "previous_value": 15,
        "change_pct": 14,
        "direction": 13,
        "relevance": 16,
        "commercial_interpretation": 60,
        "compared_against_timestamp": 22,
    }

    for name, width in widths.items():
        col_idx = h.get(name)
        if col_idx:
            col = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col].width = width


def align_text_columns(ws) -> None:
    h = headers(ws)

    centered = {
        "platform",
        "account_name",
        "creative_group",
        "metric",
        "direction",
        "relevance",
        "health_status",
        "client_number",
    }

    for name, col_idx in h.items():
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if name in centered:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif name in TEXT_COLUMNS or "description" in name or "interpretation" in name:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_sheet(ws) -> None:
    if ws.max_row < 1:
        return

    restore_timestamp_header_if_needed(ws)
    apply_tab_color(ws)
    hide_technical_columns(ws)
    style_headers(ws)
    apply_body_style(ws)
    apply_bold_columns(ws)
    apply_number_formats(ws)
    color_health_status(ws)
    color_relevance_direction_severity(ws)
    set_widths(ws)
    align_text_columns(ws)


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"No existe: {MASTER_PATH}")
        return 1

    try:
        backup()
        wb = load_workbook(MASTER_PATH)

        for ws in wb.worksheets:
            style_sheet(ws)

        wb.save(MASTER_PATH)
        print("Estilo final aplicado: columnas técnicas ocultas, estética sobria y formatos corregidos.")
        return 0

    except PermissionError:
        print("No se pudo guardar. Cerrá Excel y volvé a ejecutar.")
        return 2

    except Exception as exc:
        print(f"ERROR aplicando estilo final: {exc}")
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
