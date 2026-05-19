from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

MASTER = Path(r"D:\Proyectos\hma-system\historico\HMA_Master.xlsx")
SHEET = "recommendations"

SEVERITY_MAP = {
    "critical": "crítico",
    "critico": "crítico",
    "crítico": "crítico",
    "high": "alto",
    "alta": "alto",
    "alto": "alto",
    "opportunity": "oportunidad",
    "oportunidad": "oportunidad",
    "warning": "advertencia",
    "advertencia": "advertencia",
    "info": "informativo",
    "informativo": "informativo",
}

TYPE_MAP = {
    "alert": "alerta",
    "alerta": "alerta",
    "opportunity": "oportunidad",
    "oportunidad": "oportunidad",
    "warning": "advertencia",
    "advertencia": "advertencia",
    "info": "informativo",
    "informativo": "informativo",
}

PATTERN_MAP = {
    "CPA up + ROAS down": "CPA sube + ROAS baja",
    "ROAS down + revenue down": "ROAS baja + ingresos bajan",
    "CTR up + CVR down": "CTR sube + CVR baja",
    "CVR down + CPA up": "CVR baja + CPA sube",
    "CTR down + CPC up": "CTR baja + CPC sube",
    "Spend up + conversions not following": "Inversión sube + conversiones no acompañan",
    "Comparison gap > 1 hour": "Brecha de comparación mayor a 1 hora",
    "ROAS up + CPA down + conversions up": "ROAS sube + CPA baja + conversiones suben",
}

def clean(v):
    return "" if v is None else str(v).strip()

def key(v):
    return clean(v).lower()

def header_map(ws):
    return {key(c.value): c.column for c in ws[1] if c.value}

def get(ws, h, row, name):
    col = h.get(name)
    return ws.cell(row, col).value if col else None

def setv(ws, h, row, name, value):
    col = h.get(name)
    if col:
        ws.cell(row, col).value = value

def summarize(rows):
    severities = Counter(r["severidad"] for r in rows)
    patterns = list(dict.fromkeys([r["patron"] for r in rows if r["patron"]]))[:4]

    crit = severities.get("crítico", 0)
    high = severities.get("alto", 0)
    opp = severities.get("oportunidad", 0)
    warn = severities.get("advertencia", 0)

    parts = []
    if crit:
        parts.append(f"{crit} crítica(s)")
    if high:
        parts.append(f"{high} alta(s)")
    if opp:
        parts.append(f"{opp} oportunidad(es)")
    if warn:
        parts.append(f"{warn} advertencia(s)")

    if parts:
        text = "Hora con " + ", ".join(parts)
    else:
        text = "Hora sin señales fuertes"

    if patterns:
        text += ": " + "; ".join(patterns) + "."

    if crit:
        text += " Acción central: no escalar; auditar rentabilidad, CPA, ROAS, tracking y calidad de conversión."
    elif high:
        text += " Acción central: revisar hoy antes de mover presupuesto."
    elif opp:
        text += " Acción central: evaluar escala controlada con validación de estabilidad."
    elif warn:
        text += " Acción central: revisar calidad de datos antes de decidir."
    else:
        text += " Acción central: monitorear."

    return text

wb = load_workbook(MASTER)

if SHEET not in wb.sheetnames:
    raise SystemExit("No existe la hoja recommendations.")

ws = wb[SHEET]
h = header_map(ws)

required = ["fecha_hora", "cliente", "resumen_horario", "prioridad", "tipo", "severidad", "area", "patron_detectado"]
missing = [x for x in required if x not in h]

if missing:
    raise SystemExit(f"Faltan columnas esperadas en recommendations: {missing}")

# 1) Normalizar valores inglés/español.
for row in range(2, ws.max_row + 1):
    sev_raw = clean(get(ws, h, row, "severidad"))
    tipo_raw = clean(get(ws, h, row, "tipo"))
    patron_raw = clean(get(ws, h, row, "patron_detectado"))

    sev = SEVERITY_MAP.get(key(sev_raw), sev_raw)
    tipo = TYPE_MAP.get(key(tipo_raw), tipo_raw)
    patron = PATTERN_MAP.get(patron_raw, patron_raw)

    setv(ws, h, row, "severidad", sev)
    setv(ws, h, row, "tipo", tipo)
    setv(ws, h, row, "patron_detectado", patron)

# 2) Reconstruir resumen_horario por fecha_hora + cliente.
groups = defaultdict(list)

for row in range(2, ws.max_row + 1):
    fecha = clean(get(ws, h, row, "fecha_hora"))
    cliente = clean(get(ws, h, row, "cliente"))

    if not fecha:
        continue

    groups[(fecha, cliente)].append({
        "row": row,
        "severidad": clean(get(ws, h, row, "severidad")),
        "patron": clean(get(ws, h, row, "patron_detectado")),
    })

for _, rows in groups.items():
    resumen = summarize(rows)
    for item in rows:
        setv(ws, h, item["row"], "resumen_horario", resumen)

# 3) Estilo visual.
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
thin = Side(style="thin", color="E8EEF7")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=thin, right=thin)

widths = {
    "fecha_hora": 22,
    "cliente": 20,
    "resumen_horario": 82,
    "prioridad": 24,
    "tipo": 16,
    "severidad": 16,
    "area": 22,
    "patron_detectado": 38,
    "accion_recomendada": 72,
    "accion_bloqueada": 62,
    "motivo": 82,
    "confianza": 16,
    "requiere_revision_humana": 24,
    "calidad_comparacion": 28,
    "brecha_comparacion_horas": 24,
    "estado_muestra": 20,
    "metricas_fuente": 34,
}

for header, col in h.items():
    if header in widths:
        ws.column_dimensions[get_column_letter(col)].width = widths[header]

severity_colors = {
    "crítico": ("FEE2E2", "991B1B"),
    "alto": ("FEF3C7", "92400E"),
    "oportunidad": ("DCFCE7", "166534"),
    "advertencia": ("E0F2FE", "075985"),
    "informativo": ("F1F5F9", "334155"),
}

type_colors = {
    "alerta": ("FEE2E2", "991B1B"),
    "oportunidad": ("DCFCE7", "166534"),
    "advertencia": ("E0F2FE", "075985"),
    "informativo": ("F1F5F9", "334155"),
}

for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 58

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Calibri", size=10, color="111827")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)

    # prioridad
    pcol = h.get("prioridad")
    if pcol:
        cell = ws.cell(row, pcol)
        cell.font = Font(name="Calibri", size=10, bold=True, color="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # tipo
    tcol = h.get("tipo")
    if tcol:
        cell = ws.cell(row, tcol)
        fill, font = type_colors.get(key(cell.value), ("F8FAFC", "334155"))
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=10, bold=True, color=font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # severidad
    scol = h.get("severidad")
    if scol:
        cell = ws.cell(row, scol)
        fill, font = severity_colors.get(key(cell.value), ("F8FAFC", "334155"))
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=10, bold=True, color=font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # resumen horario con fondo suave
    rcol = h.get("resumen_horario")
    if rcol:
        cell = ws.cell(row, rcol)
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.font = Font(name="Calibri", size=10, bold=False, color="111827")

# 4) Freeze, filtro y tab.
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.sheet_properties.tabColor = "6A1B9A"

wb.save(MASTER)

print("recommendations formateada y resumen_horario corregido.")
print("Columnas:", [c.value for c in ws[1]])
print("Filas:", ws.max_row - 1)
