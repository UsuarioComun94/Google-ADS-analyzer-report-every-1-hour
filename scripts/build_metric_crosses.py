from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(r"D:\Proyectos\hma-system")
MASTER_PATH = BASE_DIR / "historico" / "HMA_Master.xlsx"

SOURCE_SHEET = "metric_comparison"
TARGET_SHEET = "metric_crosses"

HEADERS = [
    "timestamp",
    "client_number",
    "cross_id",
    "pattern",
    "severity",
    "confidence",
    "sample_status",
    "comparison_gap_hours",
    "comparison_quality",
    "likely_bottleneck",
    "evidence",
    "recommended_action",
    "blocked_action",
    "required_checks",
    "source_metrics",
]


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def lower(value: Any) -> str:
    return norm(value).lower()


def parse_float(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = str(value).strip()
    text = (
        text.replace("$", "")
        .replace("%", "")
        .replace("x", "")
        .replace("X", "")
        .replace(" ", "")
    )

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None


def parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value

    text = norm(value)
    if not text:
        return None

    text = text.replace("T", " ")

    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]:
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            pass

    return None


def headers(ws) -> dict[str, int]:
    result = {}
    for cell in ws[1]:
        key = lower(cell.value)
        if key:
            result[key] = cell.column
    return result


def group_metric_comparison(ws) -> dict[tuple[str, str], dict[str, Any]]:
    h = headers(ws)

    ts_col = h.get("timestamp")
    client_col = h.get("client_number")
    compared_ts_col = h.get("compared_against_timestamp")
    metric_col = h.get("metric")
    current_col = h.get("current_value")
    previous_col = h.get("previous_value")
    change_col = h.get("change_pct")
    relevance_col = h.get("relevance")

    required = {
        "timestamp": ts_col,
        "metric": metric_col,
        "current_value": current_col,
        "previous_value": previous_col,
        "change_pct": change_col,
    }

    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise RuntimeError(f"Faltan columnas obligatorias en metric_comparison: {missing}")

    groups: dict[tuple[str, str], dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        timestamp_raw = ws.cell(row=row, column=ts_col).value
        timestamp = norm(timestamp_raw)

        if not timestamp:
            continue

        client = norm(ws.cell(row=row, column=client_col).value) if client_col else "CLIENTE-DEMO-0001"
        metric_name = lower(ws.cell(row=row, column=metric_col).value)

        if not metric_name:
            continue

        key = (timestamp, client)

        if key not in groups:
            compared_raw = ws.cell(row=row, column=compared_ts_col).value if compared_ts_col else None
            ts_dt = parse_dt(timestamp_raw)
            compared_dt = parse_dt(compared_raw)

            gap_hours = None
            if ts_dt and compared_dt:
                gap_hours = round((ts_dt - compared_dt).total_seconds() / 3600, 2)

            groups[key] = {
                "timestamp": timestamp,
                "client_number": client,
                "comparison_gap_hours": gap_hours,
                "metrics": {},
            }

        groups[key]["metrics"][metric_name] = {
            "current": parse_float(ws.cell(row=row, column=current_col).value),
            "previous": parse_float(ws.cell(row=row, column=previous_col).value),
            "change": parse_float(ws.cell(row=row, column=change_col).value),
            "relevance": lower(ws.cell(row=row, column=relevance_col).value) if relevance_col else "",
        }

    return groups


def current(metrics: dict[str, Any], name: str) -> float | None:
    item = metrics.get(name)
    return None if not item else item.get("current")


def change(metrics: dict[str, Any], name: str) -> float | None:
    item = metrics.get(name)
    return None if not item else item.get("change")


def fmt_pct(value: float | None) -> str:
    if value is None:
        return "s/d"
    return f"{value:.1f}%"


def fmt_money(value: float | None) -> str:
    if value is None:
        return "s/d"
    return f"${value:,.2f}"


def fmt_num(value: float | None) -> str:
    if value is None:
        return "s/d"
    return f"{value:,.0f}"


def sample_status(metrics: dict[str, Any]) -> str:
    clicks = current(metrics, "clicks")
    conversions = current(metrics, "conversions")

    if clicks is None and conversions is None:
        return "unknown"

    if clicks is not None and clicks < 50:
        return "low_clicks"

    if conversions is not None and conversions < 5:
        return "low_conversions"

    return "valid"


def comparison_quality(group: dict[str, Any]) -> str:
    gap = group.get("comparison_gap_hours")

    if gap is None:
        return "unknown"

    if gap <= 1.5:
        return "clean_hourly_comparison"

    return "degraded_missing_hours"


def adjust_confidence(base: str, group: dict[str, Any]) -> str:
    sample = sample_status(group["metrics"])
    quality = comparison_quality(group)

    if sample != "valid" or quality != "clean_hourly_comparison":
        if base == "high":
            return "medium"
        if base == "medium/high":
            return "medium"
        return "low/medium"

    return base


def add_cross(
    rows: list[dict[str, Any]],
    group: dict[str, Any],
    cross_id: str,
    pattern: str,
    severity: str,
    confidence: str,
    likely_bottleneck: str,
    evidence: str,
    recommended_action: str,
    blocked_action: str,
    required_checks: str,
    source_metrics: str,
) -> None:
    rows.append(
        {
            "timestamp": group["timestamp"],
            "client_number": group["client_number"],
            "cross_id": cross_id,
            "pattern": pattern,
            "severity": severity,
            "confidence": adjust_confidence(confidence, group),
            "sample_status": sample_status(group["metrics"]),
            "comparison_gap_hours": group.get("comparison_gap_hours"),
            "comparison_quality": comparison_quality(group),
            "likely_bottleneck": likely_bottleneck,
            "evidence": evidence,
            "recommended_action": recommended_action,
            "blocked_action": blocked_action,
            "required_checks": required_checks,
            "source_metrics": source_metrics,
        }
    )


def build_crosses(groups: dict[tuple[str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for group in groups.values():
        m = group["metrics"]

        spend_ch = change(m, "spend")
        impressions_ch = change(m, "impressions")
        clicks_ch = change(m, "clicks")
        conversions_ch = change(m, "conversions")
        revenue_ch = change(m, "revenue")
        ctr_ch = change(m, "ctr")
        cpc_ch = change(m, "cpc")
        cvr_ch = change(m, "cvr")
        cpa_ch = change(m, "cpa")
        roas_ch = change(m, "roas")

        clicks_now = current(m, "clicks")
        conversions_now = current(m, "conversions")
        ctr_now = current(m, "ctr")
        cpc_now = current(m, "cpc")
        cvr_now = current(m, "cvr")
        cpa_now = current(m, "cpa")
        roas_now = current(m, "roas")
        spend_now = current(m, "spend")

        # 1. CTR alto/sube + CVR bajo/cae
        if ctr_ch is not None and cvr_ch is not None and ctr_ch >= 8 and cvr_ch <= -10:
            add_cross(
                rows,
                group,
                "ctr_up_cvr_down",
                "CTR up + CVR down",
                "high",
                "medium/high",
                "landing|offer|form|message_match|tracking",
                f"CTR cambió {fmt_pct(ctr_ch)} y CVR cambió {fmt_pct(cvr_ch)}. CTR actual {fmt_pct(ctr_now)}, CVR actual {fmt_pct(cvr_now)}.",
                "Auditar landing, oferta, formulario, velocidad, congruencia anuncio-página y eventos de conversión.",
                "No escalar presupuesto. No pausar campaña antes de validar funnel post-click.",
                "tracking_health|sample_status|comparison_gap|rolling_average",
                "ctr,cvr",
            )

        # 2. CTR baja + CPC sube
        if ctr_ch is not None and cpc_ch is not None and ctr_ch <= -10 and cpc_ch >= 10:
            add_cross(
                rows,
                group,
                "ctr_down_cpc_up",
                "CTR down + CPC up",
                "high",
                "high",
                "creative|relevance|auction_pressure|keyword_match|audience",
                f"CTR cambió {fmt_pct(ctr_ch)} y CPC cambió {fmt_pct(cpc_ch)}. CPC actual {fmt_money(cpc_now)}.",
                "Revisar copy, títulos, creativos, términos de búsqueda, segmentación, pujas y relevancia.",
                "No aumentar pujas ni presupuesto hasta resolver relevancia/costo de click.",
                "sample_status|comparison_gap|auction_context",
                "ctr,cpc",
            )

        # 3. Clicks suben + conversiones bajan o quedan planas
        if clicks_ch is not None and conversions_ch is not None and clicks_ch >= 15 and conversions_ch <= 0:
            add_cross(
                rows,
                group,
                "clicks_up_conversions_flat_or_down",
                "Clicks up + conversions flat/down",
                "high",
                "medium/high",
                "traffic_quality|intent|landing|tracking",
                f"Clicks cambiaron {fmt_pct(clicks_ch)} y conversiones cambiaron {fmt_pct(conversions_ch)}. Clicks actuales {fmt_num(clicks_now)}, conversiones actuales {fmt_num(conversions_now)}.",
                "Revisar calidad de tráfico, intención, términos/placements, landing y tracking de conversiones.",
                "No optimizar por clicks. No escalar solo porque aumentó el volumen.",
                "tracking_health|sample_status|comparison_gap",
                "clicks,conversions",
            )

        # 4. Spend sube + conversiones no acompañan
        if spend_ch is not None and spend_ch >= 20 and (conversions_ch is None or conversions_ch <= 5):
            add_cross(
                rows,
                group,
                "spend_up_conversions_not_following",
                "Spend up + conversions not following",
                "high",
                "medium/high",
                "budget_pacing|traffic_quality|conversion_efficiency",
                f"Spend cambió {fmt_pct(spend_ch)} y conversiones cambiaron {fmt_pct(conversions_ch)}. Spend actual {fmt_money(spend_now)}.",
                "Revisar pacing, distribución de presupuesto, segmentos que absorbieron gasto y eficiencia por campaña.",
                "No aumentar presupuesto general. No reforzar campañas que gastan más sin producir más conversiones.",
                "sample_status|comparison_gap|budget_pacing",
                "spend,conversions",
            )

        # 5. CPA sube + ROAS baja
        if cpa_ch is not None and roas_ch is not None and cpa_ch >= 15 and roas_ch <= -15:
            add_cross(
                rows,
                group,
                "cpa_up_roas_down",
                "CPA up + ROAS down",
                "critical",
                "high",
                "profitability|conversion_quality|budget_efficiency",
                f"CPA cambió {fmt_pct(cpa_ch)} y ROAS cambió {fmt_pct(roas_ch)}. CPA actual {fmt_money(cpa_now)}, ROAS actual {roas_now if roas_now is not None else 's/d'}x.",
                "Auditar rentabilidad, costo de adquisición, valor de conversión, campañas con mayor gasto y calidad de conversiones.",
                "No escalar. No aumentar pujas. No declarar performance positiva aunque haya más tráfico.",
                "tracking_health|sample_status|comparison_gap|revenue_validation",
                "cpa,roas",
            )

        # 6. ROAS baja + revenue baja
        if roas_ch is not None and revenue_ch is not None and roas_ch <= -15 and revenue_ch <= -15:
            add_cross(
                rows,
                group,
                "roas_down_revenue_down",
                "ROAS down + revenue down",
                "critical",
                "high",
                "revenue_quality|conversion_value|commercial_value",
                f"ROAS cambió {fmt_pct(roas_ch)} y revenue cambió {fmt_pct(revenue_ch)}.",
                "Revisar valor de conversión, revenue real, mix de campañas, tracking de valor y calidad comercial.",
                "No optimizar solo por conversiones si el valor generado cae.",
                "revenue_validation|tracking_health|comparison_gap",
                "roas,revenue",
            )

        # 7. CVR baja + CPA sube
        if cvr_ch is not None and cpa_ch is not None and cvr_ch <= -15 and cpa_ch >= 15:
            add_cross(
                rows,
                group,
                "cvr_down_cpa_up",
                "CVR down + CPA up",
                "high",
                "high",
                "landing|offer|form|traffic_quality|tracking",
                f"CVR cambió {fmt_pct(cvr_ch)} y CPA cambió {fmt_pct(cpa_ch)}. CVR actual {fmt_pct(cvr_now)}, CPA actual {fmt_money(cpa_now)}.",
                "Revisar landing, formulario, oferta, tracking, calidad del tráfico y cambios recientes en el funnel.",
                "No escalar presupuesto ni juzgar creatividad antes de aislar el problema de conversión.",
                "tracking_health|sample_status|comparison_gap",
                "cvr,cpa",
            )

        # 8. CPC baja + CVR estable/sube: mejora potencial de eficiencia
        if cpc_ch is not None and cvr_ch is not None and cpc_ch <= -10 and cvr_ch >= -5:
            add_cross(
                rows,
                group,
                "cpc_down_cvr_stable_or_up",
                "CPC down + CVR stable/up",
                "opportunity",
                "medium/high",
                "efficient_traffic|auction_improvement",
                f"CPC cambió {fmt_pct(cpc_ch)} y CVR cambió {fmt_pct(cvr_ch)}.",
                "Mantener monitoreo. Evaluar si el tráfico más barato mantiene calidad y si CPA/ROAS acompañan.",
                "No escalar agresivamente hasta confirmar estabilidad en 2-3 cortes horarios.",
                "sample_status|comparison_gap|rolling_average",
                "cpc,cvr",
            )

        # 9. ROAS sube + CPA baja + conversiones suben: oportunidad de escala controlada
        if (
            roas_ch is not None and roas_ch >= 20
            and cpa_ch is not None and cpa_ch <= -10
            and conversions_ch is not None and conversions_ch >= 10
        ):
            add_cross(
                rows,
                group,
                "roas_up_cpa_down_conversions_up",
                "ROAS up + CPA down + conversions up",
                "opportunity",
                "medium/high",
                "scale_candidate|efficient_growth",
                f"ROAS cambió {fmt_pct(roas_ch)}, CPA cambió {fmt_pct(cpa_ch)} y conversiones cambiaron {fmt_pct(conversions_ch)}.",
                "Evaluar aumento gradual de presupuesto o reasignación incremental hacia campañas/segmentos eficientes.",
                "No duplicar presupuesto. No hacer cambios múltiples simultáneos.",
                "sample_status|comparison_gap|rolling_average|revenue_validation",
                "roas,cpa,conversions",
            )

        # 10. Clicks altos + conversiones cero
        if clicks_now is not None and clicks_now >= 50 and conversions_now is not None and conversions_now <= 0:
            add_cross(
                rows,
                group,
                "clicks_high_conversions_zero",
                "High clicks + zero conversions",
                "critical",
                "medium/high",
                "tracking|landing|offer|traffic_intent",
                f"Clicks actuales {fmt_num(clicks_now)} y conversiones actuales {fmt_num(conversions_now)}.",
                "Bloquear conclusiones de performance hasta auditar tracking, eventos, formulario, landing y calidad de tráfico.",
                "No pausar ni escalar hasta validar si el problema es medición o conversión real.",
                "tracking_health|required_manual_validation",
                "clicks,conversions",
            )

        # 11. Gap horario: comparación degradada
        gap = group.get("comparison_gap_hours")
        if gap is not None and gap > 1.5:
            add_cross(
                rows,
                group,
                "degraded_comparison_gap",
                "Comparison gap > 1 hour",
                "warning",
                "high",
                "data_continuity|missing_hours",
                f"La comparación tiene un gap de {gap} horas.",
                "Tratar la comparación como degradada. Revisar continuidad de reportes antes de tomar decisiones fuertes.",
                "No declarar tendencia, no pausar ni escalar con base en comparación salteada.",
                "data_continuity|comparison_gap",
                "timestamp,compared_against_timestamp",
            )

        # 12. Muestra baja: bloquear decisiones fuertes
        sample = sample_status(m)
        if sample in {"low_clicks", "low_conversions", "unknown"}:
            add_cross(
                rows,
                group,
                f"sample_status_{sample}",
                "Low sample / insufficient statistical base",
                "warning",
                "high",
                "sample_size|decision_quality",
                f"Sample status: {sample}. Clicks actuales {fmt_num(clicks_now)}, conversiones actuales {fmt_num(conversions_now)}.",
                "Usar esta hora como señal preliminar. Esperar más muestra o validar con promedio móvil antes de actuar fuerte.",
                "No pausar, no escalar ni concluir tendencia por una muestra insuficiente.",
                "sample_status|rolling_average",
                "clicks,conversions",
            )

    # Deduplicación por timestamp + cliente + cross_id.
    unique: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["timestamp"], row["client_number"], row["cross_id"])
        unique[key] = row

    severity_order = {
        "critical": 0,
        "high": 1,
        "warning": 2,
        "opportunity": 3,
        "medium": 4,
        "low": 5,
    }

    return sorted(
        unique.values(),
        key=lambda r: (
            r["timestamp"],
            r["client_number"],
            severity_order.get(lower(r["severity"]), 9),
            r["cross_id"],
        ),
    )


def rebuild_sheet(wb, rows: list[dict[str, Any]]) -> None:
    if TARGET_SHEET in wb.sheetnames:
        ws = wb[TARGET_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(TARGET_SHEET)

    ws.append(HEADERS)

    if not rows:
        rows = [
            {
                "timestamp": "",
                "client_number": "",
                "cross_id": "no_metric_cross_triggered",
                "pattern": "No metric cross triggered",
                "severity": "info",
                "confidence": "medium",
                "sample_status": "",
                "comparison_gap_hours": "",
                "comparison_quality": "",
                "likely_bottleneck": "",
                "evidence": "No se detectaron cruces relevantes con los umbrales actuales.",
                "recommended_action": "Mantener monitoreo.",
                "blocked_action": "No tomar decisiones agresivas sin señales adicionales.",
                "required_checks": "",
                "source_metrics": "",
            }
        ]

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E8EEF7")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)

    widths = {
        "A": 22,
        "B": 20,
        "C": 34,
        "D": 34,
        "E": 14,
        "F": 16,
        "G": 18,
        "H": 20,
        "I": 24,
        "J": 36,
        "K": 68,
        "L": 68,
        "M": 58,
        "N": 34,
        "O": 28,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    severity_styles = {
        "critical": ("FEE2E2", "991B1B"),
        "high": ("FEF3C7", "92400E"),
        "warning": ("E0F2FE", "075985"),
        "opportunity": ("DCFCE7", "166534"),
        "info": ("F1F5F9", "334155"),
    }

    for row_idx in range(2, ws.max_row + 1):
        severity = lower(ws.cell(row=row_idx, column=5).value)
        fill_color, font_color = severity_styles.get(severity, ("F8FAFC", "334155"))

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(bottom=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10, color="111827")

        sev_cell = ws.cell(row=row_idx, column=5)
        sev_cell.fill = PatternFill("solid", fgColor=fill_color)
        sev_cell.font = Font(name="Calibri", size=10, bold=True, color=font_color)
        sev_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx in [3, 4, 6, 7, 9]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10, bold=True, color="111827")

        ws.row_dimensions[row_idx].height = 58

    ws.sheet_properties.tabColor = "B71C1C"


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"No existe: {MASTER_PATH}")
        return 1

    try:
        wb = load_workbook(MASTER_PATH)

        if SOURCE_SHEET not in wb.sheetnames:
            print(f"No existe hoja fuente: {SOURCE_SHEET}")
            return 1

        source = wb[SOURCE_SHEET]
        groups = group_metric_comparison(source)
        rows = build_crosses(groups)

        rebuild_sheet(wb, rows)
        wb.save(MASTER_PATH)

        print(f"metric_crosses construido correctamente. Filas: {len(rows)}")
        return 0

    except PermissionError:
        print("No se pudo guardar HMA_Master.xlsx. Cerrá Excel y volvé a ejecutar.")
        return 2

    except Exception as exc:
        print(f"ERROR construyendo metric_crosses: {exc}")
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
