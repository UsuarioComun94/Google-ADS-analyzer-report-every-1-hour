from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"D:\Proyectos\hma-system")
MASTER_PATH = BASE_DIR / "historico" / "HMA_Master.xlsx"

SOURCE_SHEET = "metric_comparison"
TARGET_SHEET = "recommendations"

OUTPUT_HEADERS = [
    "timestamp",
    "client_number",
    "priority",
    "severity",
    "area",
    "signal",
    "recommended_action",
    "blocked_action",
    "motive",
    "confidence",
    "requires_human_review",
    "source_metrics",
    "comparison_gap_hours",
    "sample_status",
]


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def lower(value: Any) -> str:
    return norm(value).lower()


def parse_float(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = str(value).strip()
    text = (
        text.replace("$", "")
        .replace("%", "")
        .replace("x", "")
        .replace("X", "")
        .replace(" ", "")
    )

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None


def parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value

    text = norm(value)
    if not text:
        return None

    # Soporta valores tipo 2026-05-18 20:00:00
    # y también 2026-05-18T20:00:00
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]:
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            pass

    return None


def headers(ws) -> dict[str, int]:
    result = {}
    for cell in ws[1]:
        key = lower(cell.value)
        if key:
            result[key] = cell.column
    return result


def get_latest_timestamp(ws, h: dict[str, int]) -> Any:
    ts_col = h.get("timestamp")
    if not ts_col:
        return None

    latest_value = None
    latest_dt = None

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=ts_col).value
        dt = parse_dt(value)

        if dt and (latest_dt is None or dt > latest_dt):
            latest_dt = dt
            latest_value = value

    return latest_value


def build_groups(ws, h: dict[str, int], latest_only: bool = True) -> dict[tuple[str, str], dict[str, Any]]:
    ts_col = h.get("timestamp")
    client_col = h.get("client_number")
    compared_ts_col = h.get("compared_against_timestamp")
    metric_col = h.get("metric")
    current_col = h.get("current_value")
    previous_col = h.get("previous_value")
    change_col = h.get("change_pct")
    relevance_col = h.get("relevance")

    required = [ts_col, client_col, metric_col, current_col, previous_col, change_col]
    if any(col is None for col in required):
        raise RuntimeError("metric_comparison no tiene las columnas mínimas necesarias.")

    latest_ts = get_latest_timestamp(ws, h) if latest_only else None
    latest_ts_text = norm(latest_ts)

    groups: dict[tuple[str, str], dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        timestamp = ws.cell(row=row, column=ts_col).value
        timestamp_text = norm(timestamp)

        if latest_only and timestamp_text != latest_ts_text:
            continue

        client = norm(ws.cell(row=row, column=client_col).value)
        metric = lower(ws.cell(row=row, column=metric_col).value)

        if not metric:
            continue

        key = (timestamp_text, client)

        if key not in groups:
            compared_ts = ws.cell(row=row, column=compared_ts_col).value if compared_ts_col else None
            ts_dt = parse_dt(timestamp)
            compared_dt = parse_dt(compared_ts)

            gap_hours = None
            if ts_dt and compared_dt:
                gap_hours = round((ts_dt - compared_dt).total_seconds() / 3600, 2)

            groups[key] = {
                "timestamp": timestamp_text,
                "client_number": client,
                "comparison_gap_hours": gap_hours,
                "metrics": {},
            }

        groups[key]["metrics"][metric] = {
            "current": parse_float(ws.cell(row=row, column=current_col).value),
            "previous": parse_float(ws.cell(row=row, column=previous_col).value),
            "change": parse_float(ws.cell(row=row, column=change_col).value),
            "relevance": lower(ws.cell(row=row, column=relevance_col).value) if relevance_col else "",
        }

    return groups


def metric(metrics: dict[str, Any], name: str, field: str = "change") -> float | None:
    item = metrics.get(name)
    if not item:
        return None
    return item.get(field)


def value(metrics: dict[str, Any], name: str) -> float | None:
    return metric(metrics, name, "current")


def pct(x: float | None) -> str:
    if x is None:
        return "s/d"
    return f"{x:.1f}%"


def money(x: float | None) -> str:
    if x is None:
        return "s/d"
    return f"${x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def num(x: float | None) -> str:
    if x is None:
        return "s/d"
    return f"{x:,.0f}".replace(",", ".")


def sample_status(metrics: dict[str, Any]) -> str:
    clicks = value(metrics, "clicks")
    conversions = value(metrics, "conversions")

    if clicks is None and conversions is None:
        return "unknown"

    if conversions is not None and conversions < 10:
        return "low_conversions"

    if clicks is not None and clicks < 100:
        return "low_clicks"

    return "valid"


def adjust_confidence(base: str, group: dict[str, Any]) -> str:
    sample = sample_status(group["metrics"])
    gap = group.get("comparison_gap_hours")

    if sample in {"low_conversions", "low_clicks", "unknown"}:
        if base == "ALTA":
            return "MEDIA"
        return "MEDIA/BAJA"

    if gap is not None and gap > 1.5:
        if base == "ALTA":
            return "MEDIA"
        return "MEDIA/BAJA"

    return base


def add_reco(
    out: list[dict[str, Any]],
    group: dict[str, Any],
    priority: int,
    severity: str,
    area: str,
    signal: str,
    recommended_action: str,
    blocked_action: str,
    motive: str,
    confidence: str,
    source_metrics: str,
    requires_review: str = "VERDADERO",
) -> None:
    out.append(
        {
            "timestamp": group["timestamp"],
            "client_number": group["client_number"],
            "priority": priority,
            "severity": severity,
            "area": area,
            "signal": signal,
            "recommended_action": recommended_action,
            "blocked_action": blocked_action,
            "motive": motive,
            "confidence": adjust_confidence(confidence, group),
            "requires_human_review": requires_review,
            "source_metrics": source_metrics,
            "comparison_gap_hours": group.get("comparison_gap_hours"),
            "sample_status": sample_status(group["metrics"]),
        }
    )


def build_recommendations(groups: dict[tuple[str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    recommendations: list[dict[str, Any]] = []

    for group in groups.values():
        m = group["metrics"]

        spend_ch = metric(m, "spend")
        clicks_ch = metric(m, "clicks")
        conv_ch = metric(m, "conversions")
        ctr_ch = metric(m, "ctr")
        cpc_ch = metric(m, "cpc")
        cvr_ch = metric(m, "cvr")
        cpa_ch = metric(m, "cpa")
        revenue_ch = metric(m, "revenue")
        roas_ch = metric(m, "roas")

        ctr_now = value(m, "ctr")
        cvr_now = value(m, "cvr")
        cpa_now = value(m, "cpa")
        roas_now = value(m, "roas")
        clicks_now = value(m, "clicks")
        conv_now = value(m, "conversions")
        spend_now = value(m, "spend")

        # 1. Rentabilidad deteriorada fuerte.
        if roas_ch is not None and cpa_ch is not None and roas_ch <= -20 and cpa_ch >= 20:
            add_reco(
                recommendations,
                group,
                1,
                "CRÍTICO",
                "rentabilidad",
                "ROAS_DOWN_CPA_UP",
                "Auditar presupuesto, pujas, términos/audiencias y conversión antes de invertir más. Revisar si el gasto está comprando volumen rentable o solo tráfico caro.",
                "No subir presupuesto, no ampliar segmentación y no aumentar pujas hasta validar CPA, ROAS y calidad de conversiones.",
                f"ROAS cambió {pct(roas_ch)} y CPA cambió {pct(cpa_ch)}. ROAS actual {roas_now if roas_now is not None else 's/d'}x; CPA actual {money(cpa_now)}.",
                "ALTA",
                "roas,cpa",
            )

        # 2. Atracción buena, conversión mala.
        if ctr_ch is not None and cvr_ch is not None and ctr_ch >= 10 and cvr_ch <= -15:
            add_reco(
                recommendations,
                group,
                2,
                "ALTO",
                "landing_offer_form",
                "CTR_UP_CVR_DOWN",
                "Revisar landing, oferta, formulario, velocidad, mensaje post-click y tracking. El anuncio atrae, pero el tráfico no convierte proporcionalmente.",
                "No escalar presupuesto ni duplicar campaña hasta validar el cuello de botella post-click.",
                f"CTR subió {pct(ctr_ch)}, pero CVR bajó {pct(cvr_ch)}. CTR actual {pct(ctr_now)}; CVR actual {pct(cvr_now)}.",
                "ALTA",
                "ctr,cvr",
            )

        # 3. Gasto sube sin acompañamiento de conversión/rentabilidad.
        if spend_ch is not None and spend_ch >= 20 and (
            (conv_ch is not None and conv_ch <= 0) or (roas_ch is not None and roas_ch <= 0)
        ):
            add_reco(
                recommendations,
                group,
                3,
                "ALTO",
                "presupuesto",
                "SPEND_UP_NO_EFFICIENCY",
                "Revisar pacing, distribución de presupuesto y campañas/segmentos que absorbieron el gasto. Validar si el aumento de spend produjo conversiones o revenue.",
                "No aumentar presupuesto general. No redistribuir inversión hacia el mismo patrón sin revisar CPA/ROAS.",
                f"Spend subió {pct(spend_ch)}; conversiones cambiaron {pct(conv_ch)}; ROAS cambió {pct(roas_ch)}. Spend actual {money(spend_now)}.",
                "ALTA",
                "spend,conversions,roas",
            )

        # 4. Tráfico sube, conversión no acompaña.
        if clicks_ch is not None and clicks_ch >= 20 and (
            (conv_ch is not None and conv_ch <= 0) or (cvr_ch is not None and cvr_ch <= -15)
        ):
            add_reco(
                recommendations,
                group,
                4,
                "ALTO",
                "calidad_de_trafico",
                "CLICKS_UP_CONVERSIONS_DOWN",
                "Revisar intención del tráfico, términos de búsqueda, audiencia, placement, landing y tracking. Hay volumen, pero la calidad aparente no acompaña.",
                "No optimizar solo por clicks. No escalar por volumen sin validar conversión.",
                f"Clicks subieron {pct(clicks_ch)}, conversiones cambiaron {pct(conv_ch)} y CVR cambió {pct(cvr_ch)}. Clicks actuales {num(clicks_now)}; conversiones actuales {num(conv_now)}.",
                "ALTA",
                "clicks,conversions,cvr",
            )

        # 5. CTR baja fuerte.
        if ctr_ch is not None and ctr_ch <= -20:
            add_reco(
                recommendations,
                group,
                5,
                "ALTO",
                "creatividad",
                "CTR_DOWN_STRONG",
                "Revisar títulos, descripciones, imágenes, fatiga creativa, relevancia del anuncio y match con intención de búsqueda/audiencia.",
                "No subir presupuesto sobre anuncios con caída fuerte de CTR sin test creativo o revisión de relevancia.",
                f"CTR actual {pct(ctr_now)} y cambio {pct(ctr_ch)}.",
                "ALTA",
                "ctr",
            )

        # 6. CPC sube y CTR baja: posible pérdida de relevancia o competencia.
        if cpc_ch is not None and ctr_ch is not None and cpc_ch >= 15 and ctr_ch <= -10:
            add_reco(
                recommendations,
                group,
                6,
                "ALTO",
                "eficiencia_de_click",
                "CPC_UP_CTR_DOWN",
                "Revisar calidad/relevancia del anuncio, competencia, pujas, términos de búsqueda, segmentación y score de calidad cuando aplique.",
                "No aumentar pujas hasta entender por qué el click se encareció y la respuesta cayó.",
                f"CPC subió {pct(cpc_ch)} y CTR bajó {pct(ctr_ch)}. CPC actual {money(value(m, 'cpc'))}.",
                "ALTA",
                "cpc,ctr",
            )

        # 7. CVR baja fuerte.
        if cvr_ch is not None and cvr_ch <= -25:
            add_reco(
                recommendations,
                group,
                7,
                "ALTO",
                "conversion",
                "CVR_DOWN_STRONG",
                "Auditar landing, formulario, oferta, velocidad, tracking y calidad de tráfico. El problema parece posterior al click o de intención.",
                "No pausar campaña sin revisar tracking y muestra; no escalar hasta estabilizar CVR.",
                f"CVR actual {pct(cvr_now)} y cambio {pct(cvr_ch)}.",
                "MEDIA/ALTA",
                "cvr",
            )

        # 8. Conversiones caen fuerte.
        if conv_ch is not None and conv_ch <= -25:
            add_reco(
                recommendations,
                group,
                8,
                "ALTO",
                "volumen_de_conversion",
                "CONVERSIONS_DOWN_STRONG",
                "Revisar cambios recientes, presupuesto, tracking, tráfico, landing y eventos de conversión. Priorizar confirmar si la caída es real o problema de medición.",
                "No redistribuir presupuesto ni pausar de forma agresiva sin confirmar tracking y comparación horaria.",
                f"Conversiones cambiaron {pct(conv_ch)}. Conversiones actuales {num(conv_now)}.",
                "MEDIA/ALTA",
                "conversions",
            )

        # 9. Oportunidad rentable: ROAS sube, CPA baja, conversiones suben.
        if (
            roas_ch is not None and roas_ch >= 20
            and cpa_ch is not None and cpa_ch <= -10
            and conv_ch is not None and conv_ch >= 10
        ):
            add_reco(
                recommendations,
                group,
                9,
                "OPORTUNIDAD",
                "escala_controlada",
                "ROAS_UP_CPA_DOWN_CONVERSIONS_UP",
                "Evaluar aumento controlado de presupuesto o reasignación incremental hacia el segmento/campaña que generó eficiencia, manteniendo monitoreo horario.",
                "No escalar agresivamente. Evitar cambios grandes hasta confirmar estabilidad en 2-3 cortes horarios.",
                f"ROAS subió {pct(roas_ch)}, CPA bajó {pct(cpa_ch)} y conversiones subieron {pct(conv_ch)}.",
                "MEDIA/ALTA",
                "roas,cpa,conversions",
                "VERDADERO",
            )

        # 10. Revenue cae fuerte con ROAS o conversiones deterioradas.
        if revenue_ch is not None and revenue_ch <= -25 and (
            (roas_ch is not None and roas_ch <= -10) or (conv_ch is not None and conv_ch <= -10)
        ):
            add_reco(
                recommendations,
                group,
                10,
                "ALTO",
                "valor",
                "REVENUE_DOWN_WITH_EFFICIENCY_RISK",
                "Revisar valor de conversión, mix de campañas, calidad de leads/ventas y cambios de tracking. Priorizar recuperar valor, no solo volumen.",
                "No optimizar solo por cantidad de conversiones si el valor/revenue está cayendo.",
                f"Revenue cambió {pct(revenue_ch)}, ROAS cambió {pct(roas_ch)} y conversiones cambiaron {pct(conv_ch)}.",
                "MEDIA/ALTA",
                "revenue,roas,conversions",
            )

    # Deduplicar por timestamp + cliente + signal.
    deduped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for reco in recommendations:
        key = (reco["timestamp"], reco["client_number"], reco["signal"])
        if key not in deduped or reco["priority"] < deduped[key]["priority"]:
            deduped[key] = reco

    result = sorted(
        deduped.values(),
        key=lambda r: (
            0 if r["severity"] == "CRÍTICO" else 1 if r["severity"] == "ALTO" else 2,
            r["priority"],
        ),
    )

    return result[:12]


def rebuild_sheet(wb, rows: list[dict[str, Any]]) -> None:
    if TARGET_SHEET in wb.sheetnames:
        ws = wb[TARGET_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(TARGET_SHEET)

    ws.append(OUTPUT_HEADERS)

    if not rows:
        rows = [
            {
                "timestamp": "",
                "client_number": "",
                "priority": 99,
                "severity": "SIN_ALERTA_ALTA",
                "area": "estado",
                "signal": "NO_HIGH_IMPACT_RECOMMENDATION",
                "recommended_action": "No se detectaron recomendaciones de alto impacto en el último corte horario.",
                "blocked_action": "No tomar decisiones agresivas sin nuevas señales.",
                "motive": "Las variaciones actuales no activaron cruces críticos definidos.",
                "confidence": "MEDIA",
                "requires_human_review": "FALSO",
                "source_metrics": "",
                "comparison_gap_hours": "",
                "sample_status": "",
            }
        ]

    for row in rows:
        ws.append([row.get(header, "") for header in OUTPUT_HEADERS])

    style_recommendations_sheet(ws)


def style_recommendations_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E8EEF7")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin, right=thin)

    widths = {
        "A": 22,
        "B": 20,
        "C": 10,
        "D": 16,
        "E": 22,
        "F": 32,
        "G": 72,
        "H": 60,
        "I": 56,
        "J": 16,
        "K": 22,
        "L": 28,
        "M": 18,
        "N": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    severity_colors = {
        "CRÍTICO": ("FEE2E2", "991B1B"),
        "ALTO": ("FEF3C7", "92400E"),
        "OPORTUNIDAD": ("DCFCE7", "166534"),
        "SIN_ALERTA_ALTA": ("F1F5F9", "334155"),
    }

    for row in range(2, ws.max_row + 1):
        severity = norm(ws.cell(row=row, column=4).value)
        fill_color, font_color = severity_colors.get(severity, ("F8FAFC", "334155"))

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(bottom=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10, color="111827")

        sev_cell = ws.cell(row=row, column=4)
        sev_cell.fill = PatternFill("solid", fgColor=fill_color)
        sev_cell.font = Font(name="Calibri", size=10, bold=True, color=font_color)
        sev_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        priority_cell = ws.cell(row=row, column=3)
        priority_cell.font = Font(name="Calibri", size=10, bold=True, color="111827")
        priority_cell.alignment = Alignment(horizontal="center", vertical="center")

        area_cell = ws.cell(row=row, column=5)
        area_cell.font = Font(name="Calibri", size=10, bold=True, color="111827")
        area_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[row].height = 54

    ws.sheet_properties.tabColor = "6A1B9A"


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"No existe: {MASTER_PATH}")
        return 1

    try:
        wb = load_workbook(MASTER_PATH)

        if SOURCE_SHEET not in wb.sheetnames:
            print(f"No existe hoja fuente: {SOURCE_SHEET}")
            return 1

        ws = wb[SOURCE_SHEET]
        h = headers(ws)

        groups = build_groups(ws, h, latest_only=True)
        recommendations = build_recommendations(groups)

        rebuild_sheet(wb, recommendations)
        wb.save(MASTER_PATH)

        print(f"Recommendations reconstruida con alto impacto. Filas: {len(recommendations)}")
        return 0

    except PermissionError:
        print("No se pudo guardar HMA_Master.xlsx. Cerrá Excel y volvé a ejecutar.")
        return 2

    except Exception as exc:
        print(f"ERROR construyendo recommendations: {exc}")
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
