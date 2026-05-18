from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(r"D:\Proyectos\hma-system")
MASTER_PATH = BASE_DIR / "historico" / "HMA_Master.xlsx"
BACKUP_DIR = BASE_DIR / "historico" / "_logic_backups"

SHEET_NAME = "metric_comparison"

GOOD_WHEN_UP = {
    "ctr",
    "cvr",
    "conversions",
    "revenue",
    "roas",
}

GOOD_WHEN_DOWN = {
    "cpc",
    "cpa",
}

CONTEXTUAL = {
    "spend",
    "impressions",
    "clicks",
}

MONEY_METRICS = {"spend", "cpc", "cpa", "revenue"}
PERCENT_METRICS = {"ctr", "cvr"}
INTEGER_METRICS = {"impressions", "clicks", "conversions"}

FILLS = {
    "positiva": "DCFCE7",
    "crítica": "FEE2E2",
    "moderada": "FEF3C7",
    "leve": "E0F2FE",
    "ruido normal": "F8FAFC",
    "observación": "F1F5F9",
}

FONT_COLORS = {
    "positiva": "166534",
    "crítica": "991B1B",
    "moderada": "92400E",
    "leve": "075985",
    "ruido normal": "475569",
    "observación": "334155",
}


def norm(value) -> str:
    return "" if value is None else str(value).strip()


def lower(value) -> str:
    return norm(value).lower()


def parse_float(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("$", "").replace("%", "").replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None


def headers(ws) -> dict[str, int]:
    result = {}
    for cell in ws[1]:
        key = lower(cell.value)
        if key:
            result[key] = cell.column
    return result


def backup() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dst = BACKUP_DIR / f"HMA_Master_pre_metric_logic_{stamp}.xlsx"
    shutil.copy2(MASTER_PATH, dst)
    print(f"Backup creado: {dst}")


def bad_relevance(abs_change: float) -> str:
    if abs_change >= 20:
        return "crítica"
    if abs_change >= 8:
        return "moderada"
    if abs_change >= 2:
        return "leve"
    return "ruido normal"


def classify(metric: str, change_pct: float | None) -> tuple[str, str]:
    if change_pct is None:
        return "observación", "Sin cambio porcentual calculable. Revisar datos fuente."

    abs_change = abs(change_pct)

    if abs_change < 2:
        return "ruido normal", "Variación menor al umbral. No tomar decisión por esta señal aislada."

    if metric in GOOD_WHEN_UP:
        if change_pct > 0:
            if metric == "ctr":
                return "positiva", "CTR subió: mejora de respuesta del anuncio. Validar que CVR y CPA acompañen antes de escalar."
            if metric == "cvr":
                return "positiva", "CVR subió: mejora de conversión post-click. Validar muestra y tracking antes de sacar conclusión fuerte."
            if metric == "conversions":
                return "positiva", "Conversiones subieron: señal positiva de volumen si CPA/ROAS acompañan."
            if metric == "revenue":
                return "positiva", "Revenue subió: señal positiva de valor generado si spend y ROAS acompañan."
            if metric == "roas":
                return "positiva", "ROAS subió: mejora de rentabilidad relativa."
        else:
            rel = bad_relevance(abs_change)
            if metric == "ctr":
                return rel, "CTR bajó: posible pérdida de relevancia, fatiga creativa o peor match de audiencia/búsqueda."
            if metric == "cvr":
                return rel, "CVR bajó: posible problema de intención, landing, formulario, oferta o tracking."
            if metric == "conversions":
                return rel, "Conversiones bajaron: revisar tráfico, funnel, tracking y cambios recientes."
            if metric == "revenue":
                return rel, "Revenue bajó: revisar volumen, valor de conversión y calidad de ventas/leads."
            if metric == "roas":
                return rel, "ROAS bajó: rentabilidad deteriorada."

    if metric in GOOD_WHEN_DOWN:
        if change_pct < 0:
            if metric == "cpc":
                return "positiva", "CPC bajó: mejora de costo por click si la calidad del tráfico se mantiene."
            if metric == "cpa":
                return "positiva", "CPA bajó: mejora de eficiencia de adquisición si el volumen se sostiene."
        else:
            rel = bad_relevance(abs_change)
            if metric == "cpc":
                return rel, "CPC subió: tráfico más caro. Revisar competencia, calidad del anuncio, pujas o segmentación."
            if metric == "cpa":
                return rel, "CPA subió: eficiencia deteriorada. Revisar CVR, CPC, conversiones y calidad de tráfico."

    if metric == "spend":
        if abs_change >= 20:
            return "moderada", "Spend cambió fuerte. Revisar pacing y si conversiones/revenue acompañan."
        return "observación", "Spend cambió. Señal contextual: no decidir sin CPA, ROAS y conversiones."

    if metric == "impressions":
        if abs_change >= 20:
            return "moderada", "Impresiones cambiaron fuerte. Revisar entrega, presupuesto, pujas o demanda."
        return "observación", "Cambio de impresiones. Señal de volumen, no de eficiencia."

    if metric == "clicks":
        if change_pct > 0:
            return "observación", "Clicks subieron. Es positivo solo si conversiones, CVR y CPA acompañan."
        rel = bad_relevance(abs_change)
        return rel, "Clicks bajaron: revisar entrega, CTR, presupuesto, pujas o pérdida de volumen."

    return "observación", "Cambio registrado. Métrica sin polaridad definida."


def apply_relevance_style(cell, relevance: str) -> None:
    fill = FILLS.get(relevance, FILLS["observación"])
    font_color = FONT_COLORS.get(relevance, FONT_COLORS["observación"])

    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Calibri", size=10, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fix_value_formats(ws, row_idx: int, metric: str, current_col: int | None, previous_col: int | None) -> None:
    if metric in MONEY_METRICS:
        fmt = '"$" #,##0.00;[Red]-"$" #,##0.00'
    elif metric in PERCENT_METRICS:
        fmt = '0.00"%"'
    elif metric == "roas":
        fmt = "0.00x"
    elif metric in INTEGER_METRICS:
        fmt = "#,##0"
    else:
        fmt = "#,##0.00"

    for col in [current_col, previous_col]:
        if col:
            cell = ws.cell(row=row_idx, column=col)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"No existe: {MASTER_PATH}")
        return 1

    try:
        backup()

        wb = load_workbook(MASTER_PATH)

        if SHEET_NAME not in wb.sheetnames:
            print(f"No existe la hoja {SHEET_NAME}.")
            return 1

        ws = wb[SHEET_NAME]
        h = headers(ws)

        metric_col = h.get("metric")
        change_col = h.get("change_pct")
        relevance_col = h.get("relevance")
        interpretation_col = h.get("commercial_interpretation")
        current_col = h.get("current_value")
        previous_col = h.get("previous_value")

        required = {
            "metric": metric_col,
            "change_pct": change_col,
            "relevance": relevance_col,
        }

        missing = [name for name, col in required.items() if not col]
        if missing:
            print(f"Faltan columnas obligatorias: {missing}")
            return 1

        fixed = 0

        for row_idx in range(2, ws.max_row + 1):
            metric = lower(ws.cell(row=row_idx, column=metric_col).value)
            change_pct = parse_float(ws.cell(row=row_idx, column=change_col).value)

            relevance, interpretation = classify(metric, change_pct)

            rel_cell = ws.cell(row=row_idx, column=relevance_col)
            rel_cell.value = relevance
            apply_relevance_style(rel_cell, relevance)

            if interpretation_col:
                ws.cell(row=row_idx, column=interpretation_col).value = interpretation

            fix_value_formats(ws, row_idx, metric, current_col, previous_col)

            fixed += 1

        wb.save(MASTER_PATH)
        print(f"metric_comparison corregido. Filas procesadas: {fixed}")
        return 0

    except PermissionError:
        print("No se pudo guardar. Cerrá HMA_Master.xlsx y volvé a ejecutar.")
        return 2

    except Exception as exc:
        print(f"ERROR corrigiendo lógica de métricas: {exc}")
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
